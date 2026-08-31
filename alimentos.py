"""Proceso de Alimentos: validación contra catálogo y archivos bancarios."""

from __future__ import annotations

import io
import re
import unicodedata
import zipfile
from collections import Counter
from copy import copy
from dataclasses import dataclass
from datetime import date
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import sio_tema


RFC_ORDENANTE = "IGS050707596"
DESCRIPCION = "ALIMENTOS"
CUENTA_ORIGEN = "0230584794"
ENCABEZADOS_SALIDA = [
    "Oper",
    "Clave ID",
    "Cuenta Origen",
    "Cuenta destino",
    "Importe",
    "Referencia",
    "Descripción",
    "RFC Ordenante",
    "IVA",
    "Fecha aplicación",
    "Nombre beneficiario",
]


def _normalizar(valor) -> str:
    texto = "" if valor is None else str(valor)
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"[^A-Z0-9]+", " ", texto.upper()).strip()
    return re.sub(r"\s+", " ", texto)


def _texto(valor) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def _identificador(valor) -> str:
    """Convierte cuentas/CLABE a texto sin notación científica."""
    if valor is None:
        return ""
    if isinstance(valor, bool):
        return str(valor)
    if isinstance(valor, int):
        return str(valor)
    if isinstance(valor, float):
        return str(int(valor)) if valor.is_integer() else format(valor, "f").rstrip("0").rstrip(".")
    texto = str(valor).strip().replace(" ", "")
    if re.fullmatch(r"\d+\.0+", texto):
        return texto.split(".", 1)[0]
    return texto


def _importe(valor) -> float | None:
    if valor in (None, ""):
        return None
    if isinstance(valor, (int, float, Decimal)):
        return float(valor)
    texto = str(valor).strip().replace("$", "").replace(",", "")
    texto = texto.replace("(", "-").replace(")", "")
    try:
        return float(Decimal(texto))
    except (InvalidOperation, ValueError):
        return None


def _redondear_excel(valor: float) -> float:
    """Replica ROUND(..., 2) de Excel, incluido el caso exacto de medio centavo."""
    # Excel guarda algunos importes calculados como 19319.234999999997 aunque
    # su valor decimal sea 19319.235. Primero se limpia ese ruido binario.
    limpio = Decimal(str(round(valor, 10)))
    return float(limpio.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def _firma_relleno(cell) -> tuple:
    color = cell.fill.fgColor
    return (
        cell.fill.fill_type,
        color.type,
        color.rgb if color.type == "rgb" else None,
        color.indexed if color.type == "indexed" else None,
        color.theme if color.type == "theme" else None,
        color.tint,
    )


def _relleno_de_fila(ws, fila: int, columnas: list[int]):
    celdas = [ws.cell(fila, col) for col in columnas]
    firmas = [_firma_relleno(celda) for celda in celdas]
    firma = Counter(firmas).most_common(1)[0][0]
    return copy(next(c.fill for c in celdas if _firma_relleno(c) == firma))


def _buscar_tabla(ws, requeridos: set[str], limite: int = 50):
    for fila in range(1, min(ws.max_row, limite) + 1):
        mapa = {}
        for col in range(1, ws.max_column + 1):
            nombre = _normalizar(ws.cell(fila, col).value)
            if nombre:
                mapa.setdefault(nombre, col)
        if requeridos.issubset(mapa):
            return fila, mapa
    return None, {}


def _columna(mapa: dict[str, int], *opciones: str) -> int | None:
    for opcion in opciones:
        clave = _normalizar(opcion)
        if clave in mapa:
            return mapa[clave]
    return None


@dataclass
class Operacion:
    hoja: str
    encabezado: int
    registros: list[dict]
    rellenos: dict[int, object]
    encabezados_origen: list[str]


def leer_operacion(datos: bytes) -> Operacion:
    wb = load_workbook(io.BytesIO(datos), data_only=True)
    requeridos = {"RFC", "CURP", "IMPORTE", "EMPRESA", "BANCO", "CLABE"}
    elegido = None
    for ws in wb.worksheets:
        fila, mapa = _buscar_tabla(ws, requeridos)
        tiene_nombre = "NOMBRE" in mapa or "NOMBRES" in mapa
        if fila and tiene_nombre:
            elegido = (ws, fila, mapa)
            break
    if not elegido:
        raise ValueError(
            "No encontré una tabla con RFC, CURP, nombre, importe, empresa, banco y CLABE."
        )

    ws, encabezado, mapa = elegido
    columnas = {
        "RFC": _columna(mapa, "RFC"),
        "CURP": _columna(mapa, "CURP"),
        "APELLIDO PATERNO": _columna(mapa, "APELLIDO PATERNO"),
        "APELLIDO MATERNO": _columna(mapa, "APELLIDO MATERNO"),
        "NOMBRES": _columna(mapa, "NOMBRES", "NOMBRE"),
        "IMPORTE": _columna(mapa, "IMPORTE"),
        "EMPRESA": _columna(mapa, "EMPRESA"),
        "BANCO": _columna(mapa, "BANCO"),
        "CLABE": _columna(mapa, "CLABE"),
    }
    columnas_color = sorted({c for c in columnas.values() if c})
    registros, rellenos = [], {}
    bloque = 0
    firma_anterior = object()
    vacias = 0
    for fila in range(encabezado + 1, ws.max_row + 1):
        importe = _importe(ws.cell(fila, columnas["IMPORTE"]).value)
        clabe = _identificador(ws.cell(fila, columnas["CLABE"]).value)
        nombre_base = _texto(ws.cell(fila, columnas["NOMBRES"]).value)
        if not nombre_base and importe is None and not clabe:
            vacias += 1
            if registros and vacias >= 1:
                break
            continue
        vacias = 0
        if importe is None or not nombre_base or not clabe:
            continue

        if columnas["APELLIDO PATERNO"]:
            nombre_origen = " ".join(
                x
                for x in [
                    _texto(ws.cell(fila, columnas["APELLIDO PATERNO"]).value),
                    _texto(ws.cell(fila, columnas["APELLIDO MATERNO"]).value),
                    nombre_base,
                ]
                if x
            )
        else:
            nombre_origen = nombre_base

        relleno = _relleno_de_fila(ws, fila, columnas_color)
        firma = _firma_relleno(ws.cell(fila, columnas_color[0]))
        # Se usa el relleno predominante, no sólo el de la primera celda.
        firma = (
            relleno.fill_type,
            relleno.fgColor.type,
            relleno.fgColor.rgb if relleno.fgColor.type == "rgb" else None,
            relleno.fgColor.indexed if relleno.fgColor.type == "indexed" else None,
            relleno.fgColor.theme if relleno.fgColor.type == "theme" else None,
            relleno.fgColor.tint,
        )
        if firma != firma_anterior:
            bloque += 1
            firma_anterior = firma
            rellenos[bloque] = relleno

        registros.append(
            {
                "fila_origen": fila,
                "bloque": bloque,
                "RFC": _texto(ws.cell(fila, columnas["RFC"]).value),
                "CURP": _texto(ws.cell(fila, columnas["CURP"]).value),
                "APELLIDO PATERNO": _texto(ws.cell(fila, columnas["APELLIDO PATERNO"]).value)
                if columnas["APELLIDO PATERNO"] else "",
                "APELLIDO MATERNO": _texto(ws.cell(fila, columnas["APELLIDO MATERNO"]).value)
                if columnas["APELLIDO MATERNO"] else "",
                "NOMBRES": nombre_base,
                "NOMBRE ORIGEN": nombre_origen,
                "IMPORTE ORIGINAL": importe,
                "EMPRESA": _texto(ws.cell(fila, columnas["EMPRESA"]).value),
                "BANCO": _texto(ws.cell(fila, columnas["BANCO"]).value),
                "CLABE ORIGEN": clabe,
            }
        )
    if not registros:
        raise ValueError("Encontré los encabezados, pero no registros válidos debajo de ellos.")
    return Operacion(ws.title, encabezado, registros, rellenos, list(columnas))


def leer_catalogo(datos: bytes) -> dict[str, dict]:
    # El modo read_only vuelve muy lento el acceso aleatorio por encabezados en
    # catálogos grandes; cargar la hoja normal es mucho más rápido aquí.
    wb = load_workbook(io.BytesIO(datos), data_only=True, read_only=False)
    catalogo: dict[str, dict] = {}
    for ws in wb.worksheets:
        for fila in range(1, min(ws.max_row, 60) + 1):
            mapa = {
                _normalizar(ws.cell(fila, col).value): col
                for col in range(1, ws.max_column + 1)
                if _normalizar(ws.cell(fila, col).value)
            }
            clabe_inter = _columna(mapa, "CLABE INTERBANCARIA")
            clabe_alt = _columna(mapa, "CLABE")
            nombre = _columna(mapa, "NOMBRE")
            if not nombre or not (clabe_inter or clabe_alt):
                continue
            id_col = _columna(mapa, "ID ISLAS", "ID", "CLAVE ID", "ID SINDICATO")
            cuenta_col = _columna(mapa, "NO. DE CUENTA", "NO DE CUENTA", "CUENTA")
            socio_col = _columna(mapa, "ID SOCIO", "SOCIO", "SIO NUEVO")
            for r in range(fila + 1, ws.max_row + 1):
                valores_clabe = []
                if clabe_inter:
                    valores_clabe.append(_identificador(ws.cell(r, clabe_inter).value))
                if clabe_alt:
                    valores_clabe.append(_identificador(ws.cell(r, clabe_alt).value))
                valores_clabe = [v for v in valores_clabe if v]
                if not valores_clabe:
                    continue
                registro = {
                    "NOMBRE": _texto(ws.cell(r, nombre).value),
                    "ID": _texto(ws.cell(r, id_col).value) if id_col else "",
                    "CUENTA": _identificador(ws.cell(r, cuenta_col).value) if cuenta_col else "",
                    "CLABE": valores_clabe[0],
                    "SOCIO": _texto(ws.cell(r, socio_col).value) if socio_col else "",
                    "HOJA CATALOGO": ws.title,
                }
                for clave in valores_clabe:
                    catalogo.setdefault(clave, registro)
            break
    if not catalogo:
        raise ValueError("No encontré CLABE INTERBANCARIA/CLABE y NOMBRE en la base de datos.")
    return catalogo


def comparar(operacion: Operacion, catalogo: dict[str, dict]) -> list[dict]:
    resultado = []
    for origen in operacion.registros:
        encontrado = catalogo.get(origen["CLABE ORIGEN"])
        importe_redondeado = _redondear_excel(origen["IMPORTE ORIGINAL"])
        fila = dict(origen)
        fila.update(
            {
                "NOMBRE": encontrado["NOMBRE"] if encontrado else "",
                "ID": encontrado["ID"] if encontrado else "",
                "CUENTA": encontrado["CUENTA"] if encontrado else "",
                "CLABE": encontrado["CLABE"] if encontrado else "",
                "SOCIO": encontrado["SOCIO"] if encontrado else "",
                "IMPORTE": importe_redondeado,
                "DIFERENCIA": importe_redondeado - origen["IMPORTE ORIGINAL"],
                "VALIDACIÓN": bool(encontrado and encontrado["CLABE"] == origen["CLABE ORIGEN"]),
                "ENCONTRADO": bool(encontrado),
            }
        )
        fila["ES BANORTE"] = (fila["CLABE"] or fila["CLABE ORIGEN"]).startswith("072")
        resultado.append(fila)
    return resultado


def ordenar_salida(registros: list[dict]) -> list[dict]:
    validos = [r for r in registros if r["ENCONTRADO"]]
    return sorted(validos, key=lambda r: (1 if r["ES BANORTE"] else 0, r["bloque"], r["fila_origen"]))


def filas_bancarias(registros: list[dict], cuenta_origen: str = CUENTA_ORIGEN) -> list[dict]:
    cuenta_origen = _identificador(cuenta_origen)
    filas = []
    for referencia, r in enumerate(ordenar_salida(registros), start=1):
        filas.append(
            {
                "Oper": "02" if r["ES BANORTE"] else "04",
                "Clave ID": r["ID"],
                "Cuenta Origen": cuenta_origen,
                "Cuenta destino": r["CUENTA"] if r["ES BANORTE"] else r["CLABE"],
                "Importe": r["IMPORTE"],
                "Referencia": str(referencia),
                "Descripción": DESCRIPCION,
                "RFC Ordenante": RFC_ORDENANTE,
                "IVA": 0,
                "Fecha aplicación": "",
                "Nombre beneficiario": r["NOMBRE"],
                "bloque": r["bloque"],
                "es_banorte": r["ES BANORTE"],
            }
        )
    return filas


def crear_excel_comparacion(operacion: Operacion, registros: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparación"
    origen_headers = ["RFC", "CURP", "APELLIDO PATERNO", "APELLIDO MATERNO", "NOMBRES", "IMPORTE", "EMPRESA", "BANCO", "CLABE"]
    comp_headers = ["NOMBRE", "ID", "CUENTA", "CLABE", "SOCIO", "IMPORTE", "DIFERENCIA", "VALIDACIÓN"]
    negro = PatternFill("solid", fgColor="000000")
    blanco = Font(color="FFFFFF", bold=True)
    for col, nombre in enumerate(origen_headers, start=2):
        celda = ws.cell(2, col, nombre); celda.fill = negro; celda.font = blanco
    for col, nombre in enumerate(comp_headers, start=12):
        celda = ws.cell(2, col, nombre); celda.fill = negro; celda.font = blanco
    for pos, r in enumerate(registros, start=3):
        vals_origen = [
            r["IMPORTE ORIGINAL"] if h == "IMPORTE"
            else r["CLABE ORIGEN"] if h == "CLABE"
            else r[h]
            for h in origen_headers
        ]
        for col, valor in enumerate(vals_origen, start=2):
            ws.cell(pos, col, valor).fill = copy(operacion.rellenos[r["bloque"]])
        vals_comp = [r[h] for h in comp_headers]
        for col, valor in enumerate(vals_comp, start=12):
            ws.cell(pos, col, valor)
        ws.cell(pos, 17).number_format = '$#,##0.00'
        ws.cell(pos, 18).number_format = '$#,##0.0000'
        if r["ES BANORTE"]:
            ws.cell(pos, 15).fill = PatternFill("solid", fgColor="FF66FF")
        if not r["VALIDACIÓN"]:
            ws.cell(pos, 19).font = Font(color="FF0000", bold=True)
    total = len(registros) + 4
    ws.cell(total, 6, "TOTAL DISPERSIÓN").font = Font(bold=True)
    ws.cell(total, 7, f"=SUM(G3:G{len(registros)+2})").number_format = '$#,##0.00'
    ws.cell(total, 16, "TOTAL COMPARACIÓN").font = Font(bold=True)
    ws.cell(total, 17, f"=SUM(Q3:Q{len(registros)+2})").number_format = '$#,##0.00'
    for col, ancho in {2:17,3:22,4:20,5:20,6:28,7:16,8:20,9:16,10:23,12:38,13:14,14:18,15:23,16:16,17:16,18:16,19:15}.items():
        ws.column_dimensions[get_column_letter(col)].width = ancho
    salida = io.BytesIO(); wb.save(salida); return salida.getvalue()


def crear_excel_bancario(filas: list[dict], rellenos: dict[int, object]) -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = "Alimentos"
    borde = Border(*( [Side(style="thin", color="000000")] * 4 ))
    for col, titulo in enumerate(ENCABEZADOS_SALIDA, start=1):
        c = ws.cell(1, col, titulo)
        c.fill = PatternFill("solid", fgColor="000000")
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal="center")
        c.border = borde
    # Paleta de la plantilla bancaria de referencia. Los números de bloque
    # siguen viniendo del color del archivo de operación.
    paleta = {
        1: PatternFill("solid", fgColor="FFC000"),
        2: PatternFill("solid", fgColor=Color(theme=3, tint=0.6)),
        3: PatternFill("solid", fgColor="FFFF00"),
        4: PatternFill("solid", fgColor=Color(theme=9, tint=0.4)),
    }
    for fila_excel, registro in enumerate(filas, start=2):
        for col, titulo in enumerate(ENCABEZADOS_SALIDA, start=1):
            c = ws.cell(fila_excel, col, registro[titulo])
            c.fill = copy(paleta.get(registro["bloque"], rellenos[registro["bloque"]]))
            c.border = borde
        ws.cell(fila_excel, 5).number_format = '$#,##0.00'
        for col in (1, 2, 3, 4, 6):
            ws.cell(fila_excel, col).number_format = '@'
    total = len(filas) + 2
    ws.cell(total, 4, "TOTAL").font = Font(bold=True)
    ws.cell(total, 5, f"=SUM(E2:E{len(filas)+1})").number_format = '$#,##0.00'
    anchos = [10, 14, 18, 24, 16, 13, 18, 20, 10, 19, 42]
    for col, ancho in enumerate(anchos, start=1): ws.column_dimensions[get_column_letter(col)].width = ancho
    ws.freeze_panes = "A2"
    salida = io.BytesIO(); wb.save(salida); return salida.getvalue()


def _txt_grupo(filas: list[dict]) -> bytes:
    lineas = []
    for r in filas:
        valores = [r[h] for h in ENCABEZADOS_SALIDA]
        valores[4] = f"{float(valores[4]):.2f}"
        lineas.append("\t".join(_texto(v) for v in valores))
    # Los TXT de referencia están codificados en UTF-8 y usan saltos Windows.
    return ("\r\n".join(lineas) + "\r\n").encode("utf-8")


def crear_paquete(filas: list[dict], excel: bytes, comparacion: bytes, etiqueta: str, fecha: date) -> bytes:
    fecha_nombre = fecha.strftime("%d.%m.%Y")
    etiqueta = re.sub(r"[^A-Z0-9_-]+", " ", _normalizar(etiqueta)).strip() or "ALIMENTOS"
    grupos: dict[tuple[bool, int], list[dict]] = {}
    for fila in filas:
        grupos.setdefault((fila["es_banorte"], fila["bloque"]), []).append(fila)
    salida = io.BytesIO()
    with zipfile.ZipFile(salida, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"ALIMENTOS RESGUARDOS INTERBANCARIOS {fecha_nombre} ({etiqueta}).xlsx", excel)
        zf.writestr(f"ALIMENTOS COMPARACION {fecha_nombre} ({etiqueta}).xlsx", comparacion)
        contadores = {False: 0, True: 0}
        for (es_banorte, _bloque), grupo in grupos.items():
            contadores[es_banorte] += 1
            tipo = "BANORTE" if es_banorte else "INTERBANCARIOS"
            nombre = f"ALIMENTOS RESGUARDOS {tipo} {fecha_nombre} ({etiqueta}) -{contadores[es_banorte]}.txt"
            zf.writestr(nombre, _txt_grupo(grupo))
    return salida.getvalue()


def _estilo_validacion(df: pd.DataFrame):
    return df.style.map(
        lambda valor: "color: #ff4b4b; font-weight: 700" if valor is False else "",
        subset=["VALIDACIÓN"],
    ).format({"IMPORTE": "${:,.2f}", "DIFERENCIA": "${:,.4f}"})


def vista_alimentos():
    sio_tema.encabezado(
        "Alimentos",
        "Carga el Excel de resguardos y la base de datos. Se valida cada CLABE, "
        "se respetan los bloques de color y se generan los archivos bancarios.",
    )
    col1, col2 = st.columns(2)
    with col1:
        archivo_operacion = st.file_uploader(
            "Excel de alimentos o resguardos",
            type=["xlsx", "xlsm"],
            key="alimentos_operacion",
        )
    with col2:
        archivo_catalogo = st.file_uploader(
            "Base de datos",
            type=["xlsx", "xlsm"],
            key="alimentos_catalogo",
        )
    if not archivo_operacion or not archivo_catalogo:
        st.info("Arrastra los dos archivos para iniciar la comparación.")
        return
    try:
        operacion = leer_operacion(archivo_operacion.getvalue())
        catalogo = leer_catalogo(archivo_catalogo.getvalue())
        registros = comparar(operacion, catalogo)
    except Exception as exc:
        st.error(f"No pude procesar los archivos: {exc}")
        return

    encontrados = sum(r["ENCONTRADO"] for r in registros)
    falsos = sum(not r["VALIDACIÓN"] for r in registros)
    banorte = sum(r["ES BANORTE"] and r["ENCONTRADO"] for r in registros)
    bloques = len({r["bloque"] for r in registros})
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Registros", len(registros))
    c2.metric("Coincidencias", encontrados)
    c3.metric("Banorte (072)", banorte)
    c4.metric("Bloques de color", bloques)
    if falsos:
        st.warning(f"Hay {falsos} validaciones en FALSO. Se muestran en rojo y no se envían al archivo bancario si no existe coincidencia.")
    else:
        st.success("Todas las CLABE coinciden con la base de datos.")

    tabla = pd.DataFrame(
        [{k: r[k] for k in ["NOMBRE", "ID", "CUENTA", "CLABE", "SOCIO", "IMPORTE", "DIFERENCIA", "VALIDACIÓN"]} for r in registros]
    )
    st.subheader("Comparación con la base de datos")
    st.dataframe(_estilo_validacion(tabla), use_container_width=True, hide_index=True)

    inter = [r for r in registros if r["ENCONTRADO"] and not r["ES BANORTE"]]
    ban = [r for r in registros if r["ENCONTRADO"] and r["ES BANORTE"]]
    m1, m2, m3 = st.columns(3)
    m1.metric("Importe interbancario", f"${sum(r['IMPORTE'] for r in inter):,.2f}")
    m2.metric("Importe Banorte", f"${sum(r['IMPORTE'] for r in ban):,.2f}")
    m3.metric("Total", f"${sum(r['IMPORTE'] for r in inter + ban):,.2f}")

    st.caption(f"Cuenta origen fija: {CUENTA_ORIGEN}")
    fecha_archivos = st.date_input(
        "Fecha para los nombres de archivo",
        value=date.today(),
        key="alimentos_fecha",
    )
    filas = filas_bancarias(registros)
    comparacion = crear_excel_comparacion(operacion, registros)
    excel_bancario = crear_excel_bancario(filas, operacion.rellenos)
    paquete = crear_paquete(filas, excel_bancario, comparacion, operacion.hoja, fecha_archivos)
    nombre_fecha = fecha_archivos.strftime("%d%m%Y")
    d1, d2, d3 = st.columns(3)
    d1.download_button(
        "Descargar comparación",
        comparacion,
        file_name=f"ALIMENTOS_COMPARACION_{nombre_fecha}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    d2.download_button(
        "Descargar Excel bancario",
        excel_bancario,
        file_name=f"ALIMENTOS_RESGUARDOS_{nombre_fecha}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        disabled=not filas,
    )
    d3.download_button(
        "Descargar paquete ZIP",
        paquete,
        file_name=f"ALIMENTOS_{nombre_fecha}.zip",
        mime="application/zip",
        use_container_width=True,
        disabled=not filas,
    )
