"""Lectura y escritura de los Excel de dispersión CONVENIA.

Lee el .xlsx directo del XML en vez de con openpyxl: estos archivos traen un
autofiltro que openpyxl no sabe parsear y revienta con
"Value must be either numerical or a string containing a wildcard".

De cada archivo saca dos cosas:
  - el encabezado (CLIENTE, PROVEEDOR, CUENTA BANCARIA, PERIODO, TOTAL)
  - el detalle de la dispersión (tarjeta, nombre, importe, retención, pago final)

Y al revés, `escribir_acumulado_xlsx` arma el concentrado por cliente: la
dispersión a la izquierda y el cruce contra la base a la derecha, cada uno con
su renglón de totales. Para escribir sí se usa openpyxl (el problema del
autofiltro es sólo al leer los archivos que manda el banco).
"""

import io
import re
import unicodedata
import zipfile
from xml.etree import ElementTree as ET

NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
REL = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"


def normalizar_nombre(valor: str) -> str:
    """Para comparar nombres: mayúsculas, sin acentos, solo letras y dígitos.

    Así 'CAMIONES BRONCOS DEL NORTE, S.A. DE C.V.' del Excel y
    'CAMIONESBRONCOSDEL NORTESADECV' del OCR quedan idénticos.
    """
    texto = unicodedata.normalize("NFKD", (valor or "").upper())
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return re.sub(r"[^A-Z0-9]", "", texto)


def a_numero(valor) -> float | None:
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = re.sub(r"[^\d.\-]", "", str(valor).replace(",", ""))
    try:
        return float(texto)
    except ValueError:
        return None


def _columna(ref: str) -> str:
    return re.match(r"[A-Z]+", ref).group(0)


def _leer_hojas(datos: bytes):
    """[(nombre de hoja, [fila, ...])] donde fila es {columna: valor}."""
    z = zipfile.ZipFile(io.BytesIO(datos))
    compartidas = []
    if "xl/sharedStrings.xml" in z.namelist():
        for si in ET.fromstring(z.read("xl/sharedStrings.xml")):
            compartidas.append("".join(t.text or "" for t in si.iter(NS + "t")))

    libro = ET.fromstring(z.read("xl/workbook.xml"))
    rels = {r.get("Id"): r.get("Target")
            for r in ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))}

    hojas = []
    for hoja in libro.iter(NS + "sheet"):
        destino = rels.get(hoja.get(REL + "id"), "").lstrip("/")
        if not destino:
            continue
        ruta = destino if destino.startswith("xl/") else "xl/" + destino
        if ruta not in z.namelist():
            continue
        filas = []
        for fila in ET.fromstring(z.read(ruta)).iter(NS + "row"):
            celdas = {}
            for c in fila.iter(NS + "c"):
                if c.get("t") == "inlineStr":
                    # Texto escrito dentro de la propia celda, sin pasar por la
                    # tabla de cadenas. Excel casi no lo usa, pero los
                    # exportadores sí; sin esto sus columnas de texto se caen
                    # enteras y la hoja parece vacía.
                    bloque = c.find(NS + "is")
                    if bloque is None:
                        continue
                    valor = "".join(t.text or "" for t in bloque.iter(NS + "t"))
                else:
                    v = c.find(NS + "v")
                    if v is None or v.text is None:
                        continue
                    if c.get("t") == "s":
                        indice = int(v.text)
                        valor = compartidas[indice] if indice < len(compartidas) else ""
                    else:
                        valor = v.text
                if str(valor).strip():
                    celdas[_columna(c.get("r"))] = str(valor).strip()
            filas.append(celdas)
        hojas.append((hoja.get("name"), filas))
    return hojas


def _valor_a_la_derecha(fila: dict, columna: str) -> str:
    """El primer valor no vacío que sigue a esa columna en la misma fila."""
    posteriores = sorted((c for c in fila if len(c) > len(columna) or (len(c) == len(columna) and c > columna)),
                         key=lambda c: (len(c), c))
    for c in posteriores:
        if fila[c].strip():
            return fila[c].strip()
    return ""


# Cada quien titula las columnas a su manera, así que aquí viven todas las
# variantes vistas en los Excel de dispersión. Se comparan ya normalizadas
# (mayúsculas, sin acentos ni espacios) y SIEMPRE por igualdad exacta: si se
# comparara por "contiene", IMPORTE BRUTO e IMPORTE NETO A PAGAR se pisarían
# entre sí, y son columnas distintas.
ALIAS_NOMBRE = (
    "NOMBRE", "NOMBRECOMPLETO", "NOMBRECOMPLETOTITULAR",
    "NOMBRECOMPLETODETITULAR", "NOMBREDELTITULAR", "NOMBREDELBENEFICIARIO",
    "TITULAR", "BENEFICIARIO", "NOMRE",
)
ALIAS_TARJETA = (
    "TARJETA", "NOTARJETA", "NUMTARJETA", "NUMEROTARJETA", "TARJETAMONEDERO",
    # Algunos archivos titulan la misma columna de monedero como cuenta.
    "CUENTABANCARIA",
)
ALIAS_IMPORTE = ("IMPORTE", "IMPORTEBRUTO", "BRUTO")
ALIAS_RETENCION = (
    "RETENCION", "RETENCIONES", "DESCUENTO", "DESCUENTOS", "DEMERITO", "DEMERITOS",
)
# Última columna del detalle: cada archivo la nombra distinto.
ALIAS_PAGO_FINAL = (
    "PAGOFINAL", "DEPOSITOFINAL", "IMPORTENETOAPAGAR", "IMPORTENETO",
    "NETOAPAGAR", "NETO", "TOTAL", "PAGO", "DEPOSITO",
)

# Se recorre en este orden y el primer alias que empate se queda la columna.
CAMPOS_DETALLE = (
    ("tarjeta", ALIAS_TARJETA),
    ("nombre", ALIAS_NOMBRE),
    ("importe", ALIAS_IMPORTE),
    ("retencion", ALIAS_RETENCION),
    ("pago_final", ALIAS_PAGO_FINAL),
)


def _campo_de_encabezado(valor) -> str | None:
    """Reconoce una columna aun cuando el título venga compuesto o con erratas.

    Los alias exactos siguen teniendo prioridad para no confundir, por ejemplo,
    IMPORTE BRUTO con IMPORTE NETO. La tolerancia se limita al nombre de la
    persona, donde se han recibido formatos como "NOMRE (S), APELIDO...".
    """
    titulo = normalizar_nombre(valor)
    for campo, alias in CAMPOS_DETALLE:
        if titulo in alias:
            return campo

    empieza_como_nombre = titulo.startswith("NOMBRE") or titulo.startswith("NOMRE")
    menciona_apellidos = "APELLIDO" in titulo or "APELIDO" in titulo
    if empieza_como_nombre and menciona_apellidos:
        return "nombre"
    return None
ETIQUETAS_ENCABEZADO = {
    "CLIENTE": "cliente",
    "PROVEEDOR": "proveedor",
    "CUENTABANCARIA": "cuenta_bancaria",
    "PERIODO": "periodo",
}


# Renglones de cierre: llevan números pero no son gente. Si alguno cae en la
# columna del nombre, no es un pago en resguardo, es el pie de la tabla.
ETIQUETAS_CIERRE = frozenset((
    "TOTAL", "TOTALES", "SUBTOTAL", "TOTALDISPERSION", "TOTALFACTURA",
    "TOTALOPERACION", "TOTALDEOPERACION", "MONEDERO", "IVA", "SUMA", "GRANTOTAL",
))

ETIQUETAS_TOTAL_DISPERSION = frozenset((
    "TOTALDISPERSION", "TOTALDELDISPERSION", "TOTALOPERACION", "TOTALDEOPERACION",
))
ETIQUETAS_TOTAL_OPERACION = frozenset(("TOTALOPERACION", "TOTALDEOPERACION"))


def _total_calculado(registros: list[dict]) -> float | None:
    """Suma netos cuando el archivo no guarda el resultado de su fórmula total."""
    valores = []
    for registro in registros:
        pago_final = registro.get("pago_final")
        if pago_final is not None:
            valores.append(pago_final)
            continue
        importe = registro.get("importe")
        if importe is not None:
            valores.append(importe - (registro.get("retencion") or 0.0))
    return round(sum(valores), 2) if valores else None


def _persona_sin_tarjeta(registro: dict) -> bool:
    """¿Es una persona a la que le toca dinero pero no tiene tarjeta?

    Hay que separarla de todo lo demás que vive sin tarjeta en estas hojas:
    los renglones de cierre, los títulos de bloque y los encabezados repetidos
    cuando el archivo pega varias tablas una tras otra. El filtro es que traiga
    nombre de persona y algún importe: sin dinero de por medio no hay nada que
    resguardar.
    """
    titulo = normalizar_nombre(registro["nombre"])
    if not titulo or titulo in ETIQUETAS_CIERRE or titulo in ALIAS_NOMBRE:
        return False
    return any((registro.get(campo) or 0) for campo in ("importe", "retencion", "pago_final"))


def leer_excel_convenia(datos: bytes) -> dict:
    """Encabezado + detalle del primer hoja que traiga la tabla de dispersión."""
    resultado = {
        "cliente": "", "proveedor": "", "cuenta_bancaria": "", "periodo": "",
        "total": None, "hoja": "", "detalle": [], "total_dispersion": None,
        # Gente con importe pero sin tarjeta donde depositarlo: no se puede
        # dispersar, así que se queda en resguardo. Va aparte de `detalle` para
        # no alterar lo que ya consume esa lista (cruce, carga masiva...).
        "sin_tarjeta": [],
        "avisos": [],
    }

    for nombre_hoja, filas in _leer_hojas(datos):
        fila_encabezado = None
        for indice, fila in enumerate(filas):
            campos = {_campo_de_encabezado(v) for v in fila.values()}
            # Con que traiga cómo identificar a la persona y su tarjeta basta
            # para reconocer la tabla; el resto de columnas ya se acomodan.
            if "tarjeta" in campos and "nombre" in campos:
                fila_encabezado = indice
                break
        if fila_encabezado is None:
            continue

        resultado["hoja"] = nombre_hoja

        # Encabezado: etiqueta a la izquierda, valor a la derecha.
        for fila in filas[:fila_encabezado]:
            for columna, valor in fila.items():
                clave = ETIQUETAS_ENCABEZADO.get(normalizar_nombre(valor))
                if clave and not resultado[clave]:
                    resultado[clave] = _valor_a_la_derecha(fila, columna)
                    if clave == "periodo":
                        # El total de la operación vive en esta misma fila.
                        numeros = [a_numero(v) for c, v in fila.items() if c > columna]
                        numeros = [n for n in numeros if n is not None]
                        if numeros:
                            resultado["total"] = round(max(numeros), 2)

        # Algunos formatos no usan la etiqueta PERIODO y escriben directamente
        # "Semana 34 del ..." como título sobre la tabla.
        if not resultado["periodo"]:
            for fila in filas[:fila_encabezado]:
                periodo = next(
                    (
                        str(valor).strip()
                        for valor in fila.values()
                        if normalizar_nombre(valor).startswith("SEMANA")
                    ),
                    "",
                )
                if periodo:
                    resultado["periodo"] = periodo
                    break

        # Mapa columna -> campo, a partir del renglón de encabezados.
        mapa = {}
        for columna, valor in filas[fila_encabezado].items():
            campo = _campo_de_encabezado(valor)
            # El primero que llega gana. Importa cuando la hoja trae dos tablas
            # pegadas: nos quedamos con la de la izquierda.
            if campo and campo not in mapa.values():
                mapa[columna] = campo

        columna_tarjeta = next((c for c, campo in mapa.items() if campo == "tarjeta"), None)
        columna_nombre = next((c for c, campo in mapa.items() if campo == "nombre"), None)
        tiene_pago_final = "pago_final" in mapa.values()

        for fila in filas[fila_encabezado + 1:]:
            etiquetas = {normalizar_nombre(v) for v in fila.values()}
            etiqueta_total = etiquetas & ETIQUETAS_TOTAL_DISPERSION
            if etiqueta_total:
                numeros = [a_numero(v) for v in fila.values()]
                numeros = [n for n in numeros if n is not None]
                if numeros:
                    total_encontrado = round(max(numeros), 2)
                    resultado["total_dispersion"] = total_encontrado
                    if etiqueta_total & ETIQUETAS_TOTAL_OPERACION and resultado["total"] is None:
                        resultado["total"] = total_encontrado
                break

            tarjeta = (fila.get(columna_tarjeta, "") if columna_tarjeta else "").lstrip("'").strip()
            nombre = (fila.get(columna_nombre, "") if columna_nombre else "").strip()
            if not nombre:
                continue

            registro = {"tarjeta": tarjeta, "nombre": nombre,
                        "importe": None, "retencion": None, "pago_final": None}
            for columna, campo in mapa.items():
                if campo in ("importe", "retencion", "pago_final"):
                    registro[campo] = a_numero(fila.get(columna))

            # Los formatos sencillos sólo traen IMPORTE. En ellos ese importe
            # ya es lo que se deposita; si además hubiera retención, se descuenta.
            # Cuando existe una columna explícita de PAGO FINAL no se rellena:
            # un hueco ahí significa que esa persona no se dispersa este periodo.
            if not tiene_pago_final and registro["importe"] is not None:
                registro["pago_final"] = round(
                    registro["importe"] - (registro["retencion"] or 0.0), 2
                )

            if re.fullmatch(r"\d{6,20}", tarjeta):
                resultado["detalle"].append(registro)
            elif _persona_sin_tarjeta(registro):
                # Lo que traiga la celda de tarjeta no sirve como cuenta; se
                # limpia para que nadie la confunda con uno de verdad.
                registro["tarjeta"] = ""
                resultado["sin_tarjeta"].append(registro)

        if resultado["detalle"] or resultado["sin_tarjeta"]:
            if resultado["total_dispersion"] is None:
                resultado["total_dispersion"] = _total_calculado(
                    resultado["detalle"] + resultado["sin_tarjeta"]
                )
                if resultado["total_dispersion"] is not None:
                    resultado["avisos"].append(
                        "El archivo no traía un total de dispersión reconocible; "
                        "lo calculé sumando el detalle."
                    )
            break

    if resultado["total"] is None and resultado["total_dispersion"] is not None:
        resultado["avisos"].append(
            "No encontré el total de la operación junto a PERIODO; solo el total de dispersión."
        )
    return resultado


# --------------------------------------------------------------------------- #
# Base de datos de tarjetas (STOCK LISTADO ... .xls)
# --------------------------------------------------------------------------- #

ALIAS_CUENTA = ("CUENTA", "NOCUENTA", "NUMEROCUENTA")
ALIAS_NOMBRE_BD = ("NOMBRECOMPLETO", "NOMBRE")
ALIAS_CLABE = ("CLABE",)
ALIAS_TARJETA_BD = ("TARJETA",)


def _filas_xls(datos: bytes):
    """[(hoja, [[valor, ...], ...])] para el formato .xls viejo (OLE2)."""
    import xlrd  # solo hace falta para .xls

    libro = xlrd.open_workbook(file_contents=datos)
    hojas = []
    for nombre in libro.sheet_names():
        hoja = libro.sheet_by_name(nombre)
        filas = []
        for i in range(hoja.nrows):
            fila = []
            for j in range(hoja.ncols):
                valor = hoja.cell_value(i, j)
                if isinstance(valor, float) and valor == int(valor):
                    valor = int(valor)
                fila.append(str(valor).strip())
            filas.append(fila)
        hojas.append((nombre, filas))
    return hojas


def _filas_xlsx(datos: bytes):
    hojas = []
    for nombre, filas_dict in _leer_hojas(datos):
        filas = []
        for celdas in filas_dict:
            if not celdas:
                filas.append([])
                continue
            ancho = max(_indice_columna(c) for c in celdas) + 1
            fila = [""] * ancho
            for columna, valor in celdas.items():
                fila[_indice_columna(columna)] = valor
            filas.append(fila)
        hojas.append((nombre, filas))
    return hojas


def _indice_columna(letras: str) -> int:
    n = 0
    for ch in letras:
        n = n * 26 + (ord(ch) - 64)
    return n - 1


def leer_base_tarjetas(datos: bytes, nombre_archivo: str = "") -> dict:
    """Indexa la base por número de cuenta.

    Devuelve {"por_cuenta": {cuenta: {nombre, clabe, tarjeta}}, "hoja": ..., "filas": n}.
    Busca en todas las hojas la que traiga encabezados CUENTA y NOMBRE COMPLETO,
    en vez de asumir posiciones: la hoja útil no es la primera del archivo.
    """
    if nombre_archivo.lower().endswith(".xls"):
        hojas = _filas_xls(datos)
    else:
        try:
            hojas = _filas_xlsx(datos)
        except Exception:
            hojas = _filas_xls(datos)

    mejor = {"por_cuenta": {}, "hoja": "", "filas": 0, "duplicadas": 0}

    for nombre_hoja, filas in hojas:
        encabezado = None
        for indice, fila in enumerate(filas[:30]):
            titulos = {normalizar_nombre(v) for v in fila if v}
            if titulos & set(ALIAS_CUENTA) and titulos & set(ALIAS_NOMBRE_BD):
                encabezado = indice
                break
        if encabezado is None:
            continue

        col = {}
        for j, valor in enumerate(filas[encabezado]):
            titulo = normalizar_nombre(valor)
            if titulo in ALIAS_CUENTA and "cuenta" not in col:
                col["cuenta"] = j
            elif titulo in ALIAS_NOMBRE_BD and "nombre" not in col:
                col["nombre"] = j
            elif titulo in ALIAS_CLABE and "clabe" not in col:
                col["clabe"] = j
            elif titulo in ALIAS_TARJETA_BD and "tarjeta" not in col:
                col["tarjeta"] = j
        if "cuenta" not in col or "nombre" not in col:
            continue

        por_cuenta, duplicadas = {}, 0
        for fila in filas[encabezado + 1:]:
            if len(fila) <= col["cuenta"]:
                continue
            cuenta = re.sub(r"\D", "", fila[col["cuenta"]] or "")
            nombre = (fila[col["nombre"]] if len(fila) > col["nombre"] else "").strip()
            if not cuenta or not nombre:
                continue
            if cuenta in por_cuenta:
                duplicadas += 1
                continue  # se queda la primera aparición
            por_cuenta[cuenta] = {
                "nombre": nombre,
                "clabe": (fila[col["clabe"]] if "clabe" in col and len(fila) > col["clabe"] else "").strip(),
                "tarjeta": (fila[col["tarjeta"]] if "tarjeta" in col and len(fila) > col["tarjeta"] else "").strip(),
            }

        if len(por_cuenta) > len(mejor["por_cuenta"]):
            mejor = {"por_cuenta": por_cuenta, "hoja": nombre_hoja,
                     "filas": len(por_cuenta), "duplicadas": duplicadas}

    return mejor


# --------------------------------------------------------------------------- #
# Escritura del concentrado
# --------------------------------------------------------------------------- #

AZUL_IZQUIERDA = "1F6FC4"
AZUL_DERECHA = "1F3864"
# Los resguardos no son dinero que se disperse, así que no van en azul: se
# distinguen a simple vista de las dos tablas de pago.
AMBAR_RESGUARDO = "B26A00"    # lo retenido, que se le queda al cliente
GRIS_RESGUARDO = "5B6781"     # lo que no se pudo depositar y queda en Convenia
GRIS_TOTAL = "D9D9D9"
FORMATO_MONEDA = "#,##0.00"
# Formato contable: el cero se ve como "-", igual que en el archivo que se
# llenaba a mano, en vez de un 0.00 que estorba para leer la columna.
FORMATO_MONEDA_GUION = '#,##0.00;-#,##0.00;"-"'

# Columna donde arranca cada bloque (1 = A). Entre uno y otro se deja hueco,
# igual que en el archivo que se llena a mano.
COL_IZQUIERDA = 1
COL_DERECHA = 8


def _escribe_tabla(hoja, fila, col_inicio, bloque, color):
    """Escribe encabezados, filas y el renglón de totales de un bloque.

    Devuelve el número del renglón siguiente al bloque.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    encabezados = bloque["encabezados"]
    numericas = set(bloque.get("numericas", ()))
    total_en = set(bloque.get("total_en", ()))
    con_guion = set(bloque.get("columnas_guion", ()))

    def formato(indice):
        return FORMATO_MONEDA_GUION if indice in con_guion else FORMATO_MONEDA

    relleno = PatternFill("solid", fgColor=color)
    for i, titulo in enumerate(encabezados):
        celda = hoja.cell(row=fila, column=col_inicio + i, value=titulo)
        celda.fill = relleno
        celda.font = Font(bold=True, color="FFFFFF")
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fila += 1

    sumas = {i: 0.0 for i in total_en}
    for renglon in bloque["filas"]:
        for i, valor in enumerate(renglon):
            celda = hoja.cell(row=fila, column=col_inicio + i, value=valor)
            if i in numericas:
                celda.number_format = formato(i)
                if i in sumas and isinstance(valor, (int, float)):
                    sumas[i] += valor
        fila += 1

    if not bloque["filas"]:
        return fila

    borde = Border(top=Side(style="thin"))
    relleno_total = PatternFill("solid", fgColor=GRIS_TOTAL)
    etiqueta_en = bloque.get("columna_etiqueta", 0)
    for i in range(len(encabezados)):
        celda = hoja.cell(row=fila, column=col_inicio + i)
        celda.border = borde
        celda.fill = relleno_total
        celda.font = Font(bold=True)
        if i == etiqueta_en:
            celda.value = bloque.get("etiqueta_total", "TOTAL")
        elif i in sumas:
            celda.value = round(sumas[i], 2)
            celda.number_format = formato(i)
    return fila + 1


def escribir_acumulado_xlsx(bloques) -> bytes:
    """Arma el concentrado: un bloque por cliente en una sola hoja.

    Cada bloque lleva el nombre del cliente como título, la dispersión del Excel
    a la izquierda y el cruce contra la base de tarjetas a la derecha, cada tabla
    con su renglón de totales; es el acomodo del archivo que se llenaba a mano.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    libro = Workbook()
    hoja = libro.active
    hoja.title = "acumulado"

    for ancho, columnas in ((34, "AH"), (16, "BI"), (14, "CJDKEL")):
        for letra in columnas:
            hoja.column_dimensions[letra].width = ancho

    fila = 1
    for bloque in bloques:
        titulo = hoja.cell(row=fila, column=COL_IZQUIERDA,
                           value=bloque.get("cliente") or bloque.get("excel", ""))
        titulo.font = Font(bold=True, size=12)
        if bloque.get("cliente") and bloque.get("excel"):
            hoja.cell(row=fila, column=COL_DERECHA, value=bloque["excel"]).font = Font(
                italic=True, color="808080")
        fila += 1

        fin_izquierda = _escribe_tabla(hoja, fila, COL_IZQUIERDA,
                                       bloque["dispersion"], AZUL_IZQUIERDA)
        fin_derecha = fila
        if bloque.get("cruce") and bloque["cruce"]["filas"]:
            fin_derecha = _escribe_tabla(hoja, fila, COL_DERECHA,
                                         bloque["cruce"], AZUL_DERECHA)

        # Los resguardos van debajo de la dispersión, cada uno con su título, y
        # sólo si tienen renglones: si el periodo no trajo retenciones ni gente
        # sin tarjeta, el concentrado se ve igual que siempre.
        fila = max(fin_izquierda, fin_derecha)
        for clave, color in (("resguardo_cliente", AMBAR_RESGUARDO),
                             ("resguardo_convenia", GRIS_RESGUARDO)):
            sub_bloque = bloque.get(clave)
            if not (sub_bloque and sub_bloque.get("filas")):
                continue
            fila += 1
            celda = hoja.cell(row=fila, column=COL_IZQUIERDA,
                              value=sub_bloque.get("titulo", "RESGUARDO"))
            celda.font = Font(bold=True)
            fila = _escribe_tabla(hoja, fila + 1, COL_IZQUIERDA, sub_bloque, color)

        # Dos renglones en blanco entre clientes, como en el archivo original.
        fila += 2

    memoria = io.BytesIO()
    libro.save(memoria)
    return memoria.getvalue()


# --------------------------------------------------------------------------- #
# Formato de carga masiva de transferencias (el que se sube al banco)
# --------------------------------------------------------------------------- #

HOJA_CARGA = " Transferencias Masivas"   # el espacio del principio viene en la plantilla
GRIS_CARGA = "FF434343"
FORMATO_IMPORTE_CARGA = '"$"#,##0.00_);[Red]\\("$"#,##0.00\\)'
ENCABEZADOS_CARGA = ["Cuenta destino", "Importe", "Nombre del beneficiario", "Concepto"]
ANCHOS_CARGA = {"A": 28.3, "B": 22.0, "C": 25.7, "D": 29.9}


def escribir_carga_masiva(filas, concepto: str = "") -> bytes:
    """Archivo de carga masiva del banco: cuenta, importe, beneficiario y concepto.

    `filas` es [{"cuenta": str, "importe": float, "nombre": str}]. La cuenta se
    escribe como TEXTO a propósito: son tarjetas de 11 y CLABEs de 18 dígitos que
    empiezan con cero, y como número el banco las recibiría mochas.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    libro = Workbook()
    hoja = libro.active
    hoja.title = HOJA_CARGA

    for letra, ancho in ANCHOS_CARGA.items():
        hoja.column_dimensions[letra].width = ancho

    gris = PatternFill("solid", fgColor=GRIS_CARGA)
    for i, titulo in enumerate(ENCABEZADOS_CARGA, start=1):
        celda = hoja.cell(row=1, column=i, value=titulo)
        celda.fill = gris
        celda.font = Font(bold=True, color="FFFFFF", size=11, name="IBM Plex Sans")
        celda.alignment = Alignment(horizontal="center", vertical="center")

    arial = Font(size=10, name="Arial")
    for indice, renglon in enumerate(filas):
        fila = 2 + indice
        cuenta = hoja.cell(row=fila, column=1, value=str(renglon.get("cuenta", "")).strip())
        cuenta.number_format = "@"
        cuenta.font = arial

        importe = hoja.cell(row=fila, column=2, value=renglon.get("importe"))
        importe.number_format = FORMATO_IMPORTE_CARGA
        importe.font = arial

        hoja.cell(row=fila, column=3, value=renglon.get("nombre", "")).font = arial
        hoja.cell(row=fila, column=4, value=concepto).font = arial

    memoria = io.BytesIO()
    libro.save(memoria)
    return memoria.getvalue()


# --------------------------------------------------------------------------- #
# Concentrado de operaciones (hoja "CR3 CONV")
# --------------------------------------------------------------------------- #

AZUL_CR3 = "FF002060"
FORMATO_CONTABLE = '_-"$"* #,##0.00_-;\\-"$"* #,##0.00_-;_-"$"* "-"??_-;_-@_-'

# Encabezados de la fila 4, con la columna en la que va cada uno.
ENCABEZADOS_CR3 = {
    "B": "No op",
    "C": "Operación",
    "D": "Fecha de factura",
    "E": "Factura",
    "F": "Importe Neto de Depósito",
    "G": "Importe a Repartir",
    "H": "REAL A DISPERSAR",
    "I": "DEVOLUCIONES",
    "J": "Comisión 5%",
    "K": "IVA",
    "L": "Costo Fiscal",
    "M": "Comisión Neta",
    "N": "Devolución de Puntos",
    "P": "CONVENIA",
    "Q": "GASTOS DE OPERACIÓN",
    "R": "BENEFICIO",
    "S": "TOTAL INNTEC PARA COMISIONES RR",
    "T": "COMISION 2",
    "U": "COMISION ROGELIO RIVAS ",
    "V": "COM INNTEC",
    "X": "BENEFICIO FINAL",
}

# Los encabezados de la izquierda ocupan las filas 4 a 6; los de la derecha
# sólo 4 y 5, porque debajo (fila 6) viven los porcentajes del cálculo.
COMBINA_4_6 = "BCDEFGHIJKLMNQRX"
COMBINA_4_5 = "PSTUV"

ANCHOS_CR3 = {
    "A": 3.0, "B": 10.6, "C": 51.6, "D": 13.7, "E": 11.0, "F": 15.3, "G": 15.9,
    "H": 15.9, "I": 13.0, "J": 19.1, "K": 13.4, "L": 11.3, "M": 17.7, "N": 11.3,
    "O": 1.6, "P": 15.4, "Q": 12.6, "R": 14.4, "S": 14.9, "T": 16.4, "U": 16.4,
    "V": 10.0, "W": 1.4, "X": 17.1,
}

# Fórmulas de cada renglón, tal como vienen en el archivo que se llena a mano.
# Lo único que se captura son C (cliente), G (importe a repartir) y H (real a
# dispersar); todo lo demás se calcula solo dentro de Excel.
FORMULAS_CR3 = {
    "F": "=(G{f}*1.05*1.16)",
    "I": "=+G{f}-H{f}",
    "J": "=(G{f}*5%)",
    "K": "=(G{f}+J{f})*16%",
    "L": "=(F{f}*0.125*0.3)",
    "M": "=J{f}+K{f}-L{f}",
    "N": "=(F{f}/1.16)*0",
    "P": "=+H{f}*1.02",
    "R": "=+M{f}-(G{f}*$P$6)-Q{f}",
    "S": "=R{f}*$S$6",
    "X": "=+M{f}-(G{f}*$P$6)-S{f}",
}

FILA_ENCABEZADO = 4
FILA_PARAMETROS = 6
PRIMERA_FILA_DATOS = 8
# Renglones de más con la fórmula puesta, para seguir capturando operaciones a mano.
FILAS_EXTRA = 20


def escribir_operaciones_xlsx(operaciones, cliente: str = "", ejercicio=None) -> bytes:
    """Concentrado de operaciones con las fórmulas vivas, no los resultados.

    `operaciones` es [{"operacion": nombre del cliente, "importe_a_repartir": x,
    "real_a_dispersar": y}]. Se escriben esos tres datos y las fórmulas quedan
    igual que en la plantilla, así que al abrirlo Excel recalcula todo y los
    porcentajes de la fila 6 se pueden seguir tocando a mano.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    libro = Workbook()
    hoja = libro.active
    hoja.title = "CR3 CONV"

    for letra, ancho in ANCHOS_CR3.items():
        hoja.column_dimensions[letra].width = ancho

    azul = PatternFill("solid", fgColor=AZUL_CR3)
    blanca = Font(bold=True, color="FFFFFF", size=10)
    centrado = Alignment(horizontal="center", vertical="center", wrap_text=True)
    delgado = Side(style="thin")
    marco = Border(top=delgado, bottom=delgado, left=delgado, right=delgado)

    # Encabezado del documento
    hoja["B2"] = "CLIENTE:"
    hoja["B2"].fill = azul
    hoja["B2"].font = Font(bold=True, color="FFFFFF", size=9)
    hoja["B2"].alignment = Alignment(horizontal="center", vertical="center")
    hoja.merge_cells("C2:F2")
    hoja["C2"] = cliente
    hoja["C2"].font = Font(bold=True, size=9)
    hoja["J2"] = "EJERCICIO"
    hoja["J2"].font = Font(bold=True, size=9)
    hoja["K2"] = str(ejercicio) if ejercicio is not None else ""
    hoja["K2"].font = Font(bold=True, size=9)

    for letra, titulo in ENCABEZADOS_CR3.items():
        celda = hoja[f"{letra}{FILA_ENCABEZADO}"]
        celda.value = titulo
        celda.fill = azul
        celda.font = blanca
        celda.alignment = centrado
        celda.border = marco
        if letra in COMBINA_4_6:
            hoja.merge_cells(f"{letra}4:{letra}6")
        elif letra in COMBINA_4_5:
            hoja.merge_cells(f"{letra}4:{letra}5")

    # Porcentajes del cálculo, debajo de los encabezados de la derecha.
    for letra, valor in (("P", 0.02), ("S", 0.35), ("T", 0), ("U", "TOTAL A DEVOLVER"), ("V", 0.025)):
        celda = hoja[f"{letra}{FILA_PARAMETROS}"]
        celda.value = valor
        celda.fill = azul
        celda.font = blanca
        celda.alignment = centrado
        celda.border = marco

    total_filas = len(operaciones) + FILAS_EXTRA
    for indice in range(total_filas):
        fila = PRIMERA_FILA_DATOS + indice
        operacion = operaciones[indice] if indice < len(operaciones) else None

        if operacion:
            hoja[f"C{fila}"] = operacion.get("operacion", "")
            hoja[f"C{fila}"].font = Font(size=10)
            hoja[f"G{fila}"] = operacion.get("importe_a_repartir")
            hoja[f"H{fila}"] = operacion.get("real_a_dispersar")

        for letra, plantilla in FORMULAS_CR3.items():
            hoja[f"{letra}{fila}"] = plantilla.format(f=fila)

        for letra in "DEFGHIJKLMNPQRSTUVX":
            celda = hoja[f"{letra}{fila}"]
            celda.number_format = FORMATO_CONTABLE
            celda.border = marco

    memoria = io.BytesIO()
    libro.save(memoria)
    return memoria.getvalue()
