"""Vista y validaciones del archivo de Litografía."""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from difflib import SequenceMatcher
from io import BytesIO
import re
import unicodedata
from xml.sax.saxutils import escape
from zipfile import ZIP_DEFLATED, ZipFile

import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


@dataclass
class ImporteDetectado:
    valor: Decimal | None
    referencia: str


@dataclass
class FilaResumen:
    empresa: str
    importe: ImporteDetectado
    celda_empresa: object


def _texto_normalizado(valor: object) -> str:
    texto = unicodedata.normalize("NFKD", str(valor or ""))
    texto = "".join(caracter for caracter in texto if not unicodedata.combining(caracter))
    return re.sub(r"[^a-z0-9]+", " ", texto.lower()).strip()


def _decimal(valor: object) -> Decimal | None:
    if valor is None or isinstance(valor, bool):
        return None
    if isinstance(valor, (int, float, Decimal)):
        try:
            return Decimal(str(valor))
        except InvalidOperation:
            return None

    texto = str(valor).strip().replace("$", "").replace(" ", "")
    if not texto or texto.startswith("="):
        return None
    if "," in texto and "." in texto:
        texto = texto.replace(",", "") if texto.rfind(".") > texto.rfind(",") else texto.replace(".", "").replace(",", ".")
    elif "," in texto:
        partes = texto.split(",")
        texto = "".join(partes) if len(partes[-1]) == 3 else texto.replace(",", ".")
    try:
        return Decimal(texto)
    except InvalidOperation:
        return None


def _es_suma(valor: object) -> bool:
    return _texto_normalizado(valor) == "suma"


def _valor_cercano(hoja, fila: int, columna: int) -> ImporteDetectado:
    # Primero se buscan las posiciones usuales de un total: derecha y abajo.
    desplazamientos = [(0, paso) for paso in range(1, 6)]
    desplazamientos += [(paso, 0) for paso in range(1, 6)]
    desplazamientos += [(0, -paso) for paso in range(1, 4)]
    for delta_fila, delta_columna in desplazamientos:
        nueva_fila, nueva_columna = fila + delta_fila, columna + delta_columna
        if nueva_fila < 1 or nueva_columna < 1:
            continue
        celda = hoja.cell(nueva_fila, nueva_columna)
        valor = _decimal(celda.value)
        if valor is not None:
            return ImporteDetectado(valor, f"{hoja.title}!{celda.coordinate}")
    return ImporteDetectado(None, f"{hoja.title}!{hoja.cell(fila, columna).coordinate}")


def _suma_de_hoja(hoja) -> ImporteDetectado:
    candidatos: list[ImporteDetectado] = []
    for fila in hoja.iter_rows():
        for celda in fila:
            if not isinstance(celda, MergedCell) and _es_suma(celda.value):
                candidatos.append(_valor_cercano(hoja, celda.row, celda.column))
    for candidato in candidatos:
        if candidato.valor is not None:
            return candidato
    return candidatos[0] if candidatos else ImporteDetectado(None, "No se encontró la celda Suma")


def _coincidencia_hoja(valor: object, nombre_hoja: str) -> float:
    izquierda = _texto_normalizado(valor)
    derecha = _texto_normalizado(nombre_hoja)
    if not izquierda or not derecha:
        return 0.0
    if izquierda == derecha:
        return 1.0
    if izquierda in derecha or derecha in izquierda:
        return 0.9
    return SequenceMatcher(None, izquierda, derecha).ratio()


def _filas_del_resumen(hoja_resumen) -> list[FilaResumen]:
    candidatos = [
        celda
        for fila in hoja_resumen.iter_rows()
        for celda in fila
        if not isinstance(celda, MergedCell) and _es_suma(celda.value)
    ]
    if not candidatos:
        return []

    # La celda "Suma" que funciona como encabezado es la que tiene más importes
    # debajo. Así se distingue de la fila final que también se llama "Suma".
    def puntuacion_encabezado(celda) -> int:
        limite = min(hoja_resumen.max_row, celda.row + 50)
        return sum(
            _decimal(hoja_resumen.cell(fila, celda.column).value) is not None
            for fila in range(celda.row + 1, limite + 1)
        )

    encabezado = max(candidatos, key=puntuacion_encabezado)
    filas: list[FilaResumen] = []
    limite = min(hoja_resumen.max_row, encabezado.row + 50)
    for numero_fila in range(encabezado.row + 1, limite + 1):
        celdas_izquierda = [
            hoja_resumen.cell(numero_fila, columna)
            for columna in range(1, encabezado.column)
        ]
        if any(_es_suma(celda.value) for celda in celdas_izquierda):
            break

        celda_importe = hoja_resumen.cell(numero_fila, encabezado.column)
        importe = _decimal(celda_importe.value)
        if importe is None:
            continue

        celda_empresa = next(
            (
                celda
                for celda in celdas_izquierda
                if _texto_normalizado(celda.value)
                and _decimal(celda.value) is None
                and _texto_normalizado(celda.value) not in {"empresa", "suma"}
            ),
            None,
        )
        if celda_empresa is None:
            continue
        filas.append(
            FilaResumen(
                empresa=str(celda_empresa.value).strip(),
                importe=ImporteDetectado(
                    importe, f"{hoja_resumen.title}!{celda_importe.coordinate}"
                ),
                celda_empresa=celda_empresa,
            )
        )
    return filas


def _textos_de_hoja(hoja) -> list[str]:
    return [
        str(celda.value)
        for fila in hoja.iter_rows()
        for celda in fila
        if not isinstance(celda, MergedCell)
        and isinstance(celda.value, str)
        and _texto_normalizado(celda.value)
    ]


def _relacionar_resumen(
    hoja, detalle: ImporteDetectado, filas_resumen: list[FilaResumen]
) -> FilaResumen | None:
    # Un importe igual y único es la evidencia más fuerte. Además evita que
    # nombres mencionados dentro de otra hoja provoquen una asociación incorrecta.
    if detalle.valor is not None:
        coincidencias = [
            fila
            for fila in filas_resumen
            if fila.importe.valor is not None
            and fila.importe.valor == detalle.valor
        ]
        if len(coincidencias) == 1:
            return coincidencias[0]

    textos = [hoja.title, *_textos_de_hoja(hoja)]
    mejor: tuple[float, FilaResumen] | None = None
    for fila_resumen in filas_resumen:
        puntuacion = max(
            _coincidencia_hoja(fila_resumen.empresa, texto) for texto in textos
        )
        if mejor is None or puntuacion > mejor[0]:
            mejor = (puntuacion, fila_resumen)

    # ENA, por ejemplo, se relaciona con Nahuatl porque ese nombre aparece en
    # el encabezado y en el contenido de la hoja, aunque no aparezca en la pestaña.
    if mejor is not None and mejor[0] >= 0.82:
        return mejor[1]

    return None


def analizar_excel(contenido: bytes) -> tuple[pd.DataFrame, str]:
    libro = load_workbook(BytesIO(contenido), data_only=True, read_only=False)
    hojas_resumen = [hoja for hoja in libro.worksheets if _texto_normalizado(hoja.title) == "resumen"]
    if not hojas_resumen:
        raise ValueError("El archivo no contiene una hoja llamada 'Resumen'.")

    hoja_resumen = hojas_resumen[0]
    hojas_excluidas = {"resumen", "cuentas"}
    hojas_detalle = [
        hoja
        for hoja in libro.worksheets
        if _texto_normalizado(hoja.title) not in hojas_excluidas
    ]
    if not hojas_detalle:
        raise ValueError("El archivo no contiene hojas para comparar con Resumen.")

    filas_resumen = _filas_del_resumen(hoja_resumen)
    if not filas_resumen:
        raise ValueError("No se encontraron empresas e importes bajo la columna 'Suma' de Resumen.")

    filas = []
    for hoja in hojas_detalle:
        detalle = _suma_de_hoja(hoja)
        fila_resumen = _relacionar_resumen(hoja, detalle, filas_resumen)
        resumen = (
            fila_resumen.importe
            if fila_resumen is not None
            else ImporteDetectado(None, "No se encontró la empresa en Resumen")
        )
        diferencia = None
        estado = "Falta dato"
        if resumen.valor is not None and detalle.valor is not None:
            diferencia = detalle.valor - resumen.valor
            estado = "Coincide" if diferencia == Decimal("0.00") else "No coincide"
        filas.append(
            {
                "Hoja": hoja.title,
                "Empresa en Resumen": fila_resumen.empresa if fila_resumen else None,
                "Suma en Resumen": float(resumen.valor) if resumen.valor is not None else None,
                "Suma en hoja": float(detalle.valor) if detalle.valor is not None else None,
                "Diferencia": float(diferencia) if diferencia is not None else None,
                "Estado": estado,
                "Celda Resumen": resumen.referencia,
                "Celda hoja": detalle.referencia,
            }
        )
    return pd.DataFrame(filas), hoja_resumen.title


def _valor_como_texto(valor: object, quitar_apostrofe: bool = False) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        texto = str(int(valor))
    else:
        texto = str(valor)
    texto = texto.strip()
    if quitar_apostrofe:
        texto = texto.lstrip("'").strip()
    return texto


def _limpiar_clave(valor: object) -> str:
    texto = _valor_como_texto(valor)
    return re.sub(r"[.\s]+", "", texto)


def _columnas_de_personas(hoja) -> tuple[int, dict[str, int]] | None:
    for fila in hoja.iter_rows():
        encabezados = {
            _texto_normalizado(celda.value): celda.column
            for celda in fila
            if not isinstance(celda, MergedCell) and celda.value is not None
        }
        columna_cuenta = next(
            (
                columna
                for texto, columna in encabezados.items()
                if texto.startswith("no") and ("cuenta" in texto or "clabe" in texto)
            ),
            None,
        )
        requeridas = {
            "clave": encabezados.get("clave"),
            "nombre": encabezados.get("nombre"),
            "total": encabezados.get("total a depositar"),
            "cuenta": columna_cuenta,
            "banco": encabezados.get("banco"),
        }
        if all(requeridas.values()):
            return fila[0].row, requeridas
    return None


def extraer_tabla_comparacion(contenido: bytes) -> tuple[pd.DataFrame, pd.DataFrame]:
    libro = load_workbook(BytesIO(contenido), data_only=True, read_only=False)
    hojas_excluidas = {"resumen", "cuentas"}
    registros: list[dict[str, object]] = []
    registros_efectivo: list[dict[str, object]] = []

    for hoja in libro.worksheets:
        if _texto_normalizado(hoja.title) in hojas_excluidas:
            continue
        estructura = _columnas_de_personas(hoja)
        if estructura is None:
            continue
        fila_encabezado, columnas = estructura

        for numero_fila in range(fila_encabezado + 1, hoja.max_row + 1):
            clave = _limpiar_clave(hoja.cell(numero_fila, columnas["clave"]).value)
            nombre = _valor_como_texto(hoja.cell(numero_fila, columnas["nombre"]).value)
            total = _decimal(hoja.cell(numero_fila, columnas["total"]).value)
            cuenta = _valor_como_texto(
                hoja.cell(numero_fila, columnas["cuenta"]).value,
                quitar_apostrofe=True,
            )
            banco = _valor_como_texto(hoja.cell(numero_fila, columnas["banco"]).value)

            # Excluye títulos de bloque, filas de totales y el resumen al pie.
            if not nombre or total is None or not banco:
                continue

            registro = {
                "EMPRESA": hoja.title.strip(),
                "BLOQUE": "",
                "CLAVE": clave,
                "Nombre": nombre,
                "Total a Depositar": float(total),
                "No. Cuenta": cuenta,
                "Banco": banco,
            }
            if _texto_normalizado(banco) == "efectivo":
                registros_efectivo.append(registro)
            elif clave:
                registros.append(registro)

    columnas_salida = [
        "EMPRESA",
        "BLOQUE",
        "CLAVE",
        "Nombre",
        "Total a Depositar",
        "No. Cuenta",
        "Banco",
    ]
    return (
        pd.DataFrame(registros, columns=columnas_salida),
        pd.DataFrame(registros_efectivo, columns=columnas_salida),
    )


def validar_catalogo(contenido: bytes) -> int:
    libro = load_workbook(BytesIO(contenido), data_only=True, read_only=True)
    hojas = [
        hoja for hoja in libro.worksheets if _texto_normalizado(hoja.title) == "catalogo"
    ]
    if not hojas:
        raise ValueError("El archivo no contiene una hoja llamada 'Catálogo'.")
    hoja = hojas[0]
    encabezados = {
        _texto_normalizado(celda.value) for celda in next(hoja.iter_rows(min_row=1, max_row=1))
    }
    faltantes = {
        "empresa",
        "clave",
        "nombre",
        "clabe interbancaria",
        "banco",
        "id islas",
        "id sind",
    } - encabezados
    if faltantes:
        raise ValueError(
            "A la hoja Catálogo le faltan encabezados requeridos: "
            + ", ".join(sorted(faltantes))
        )
    return max(hoja.max_row - 1, 0)


def detectar_cuentas_nuevas(
    tabla_origen: pd.DataFrame, contenido_catalogo: bytes
) -> pd.DataFrame:
    libro = load_workbook(BytesIO(contenido_catalogo), data_only=True, read_only=True)
    hoja = next(
        hoja
        for hoja in libro.worksheets
        if _texto_normalizado(hoja.title) == "catalogo"
    )
    encabezados: dict[str, int] = {}
    for celda in next(hoja.iter_rows(min_row=1, max_row=1)):
        encabezado = _texto_normalizado(celda.value)
        if encabezado:
            encabezados.setdefault(encabezado, celda.column)
    columna_clave = encabezados["clave"]
    columna_clabe = encabezados["clabe interbancaria"]
    claves_catalogo: set[str] = set()
    clabes_catalogo: set[str] = set()
    for valores in hoja.iter_rows(min_row=2, values_only=True):
        clave = _limpiar_clave(valores[columna_clave - 1])
        clabe = _normalizar_cuenta(valores[columna_clabe - 1])
        if clave:
            claves_catalogo.add(clave.casefold())
        if clabe:
            clabes_catalogo.add(clabe)

    filas: list[dict[str, object]] = []
    for registro in tabla_origen.to_dict("records"):
        clave = _limpiar_clave(registro["CLAVE"])
        if _texto_normalizado(clave) == "s n":
            continue
        clabe = _normalizar_cuenta(registro["No. Cuenta"])
        motivos: list[str] = []
        if clave.casefold() not in claves_catalogo:
            motivos.append("Clave no encontrada")
        if clabe not in clabes_catalogo:
            motivos.append("CLABE no encontrada")
        if not motivos:
            continue
        filas.append(
            {
                "Empresa": registro["EMPRESA"],
                "Clave": clave,
                "Nombre": registro["Nombre"],
                "CLABE interbancaria": clabe,
                "Banco": registro["Banco"],
                "Importe": registro["Total a Depositar"],
                "Motivo": " y ".join(motivos),
            }
        )
    return pd.DataFrame(
        filas,
        columns=[
            "Empresa", "Clave", "Nombre", "CLABE interbancaria", "Banco",
            "Importe", "Motivo",
        ],
    )


def exportar_cuentas_nuevas_excel(tabla: pd.DataFrame) -> bytes:
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Cuentas nuevas"
    encabezados = list(tabla.columns)
    relleno = PatternFill("solid", fgColor="FFC00000")
    borde = Border(
        left=Side(style="thin", color="FF808080"),
        right=Side(style="thin", color="FF808080"),
        top=Side(style="thin", color="FF808080"),
        bottom=Side(style="thin", color="FF808080"),
    )
    for columna, encabezado in enumerate(encabezados, 1):
        celda = hoja.cell(1, columna, encabezado)
        celda.font = Font(name="Arial", size=9, bold=True, color="FFFFFFFF")
        celda.fill = relleno
        celda.alignment = Alignment(horizontal="center", vertical="center")
        celda.border = borde
    for fila_excel, valores in enumerate(tabla.itertuples(index=False, name=None), 2):
        for columna, valor in enumerate(valores, 1):
            if pd.isna(valor):
                valor = None
            celda = hoja.cell(fila_excel, columna, valor)
            celda.font = Font(name="Arial", size=8)
            celda.border = borde
        hoja.cell(fila_excel, 2).number_format = "@"
        hoja.cell(fila_excel, 4).number_format = "@"
        hoja.cell(fila_excel, 6).number_format = '$#,##0.00'
    for letra, ancho in {
        "A": 13, "B": 13, "C": 40, "D": 22, "E": 16, "F": 14, "G": 35,
    }.items():
        hoja.column_dimensions[letra].width = ancho
    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = f"A1:G{max(len(tabla) + 1, 1)}"
    salida = BytesIO()
    libro.save(salida)
    return salida.getvalue()


def _normalizar_cuenta(valor: object) -> str:
    return re.sub(r"[\s']+", "", _valor_como_texto(valor))


def _texto_igual_excel(izquierda: object, derecha: object) -> bool:
    if izquierda is None or derecha is None:
        return False
    return _valor_como_texto(izquierda).casefold() == _valor_como_texto(derecha).casefold()


def crear_tabla_con_catalogo(
    tabla_origen: pd.DataFrame, contenido_catalogo: bytes
) -> tuple[pd.DataFrame, pd.DataFrame, int]:
    libro = load_workbook(BytesIO(contenido_catalogo), data_only=True, read_only=True)
    hoja = next(
        hoja
        for hoja in libro.worksheets
        if _texto_normalizado(hoja.title) == "catalogo"
    )
    encabezados: dict[str, int] = {}
    for celda in next(hoja.iter_rows(min_row=1, max_row=1)):
        texto = _texto_normalizado(celda.value)
        if texto:
            # BUSCARV usa las primeras columnas del Catálogo (B:H). Si un
            # encabezado se repite más adelante, se conserva el primero.
            encabezados.setdefault(texto, celda.column)
    columnas = {
        "clave": encabezados["clave"],
        "nombre": encabezados["nombre"],
        "clabe": encabezados["clabe interbancaria"],
        "banco": encabezados["banco"],
        "id_islas": encabezados["id islas"],
        "id_sindicato": encabezados["id sind"],
    }

    # Conserva la primera coincidencia, igual que BUSCARV con coincidencia exacta.
    catalogo_por_clave: dict[str, dict[str, object]] = {}
    catalogo_por_cuenta: dict[str, dict[str, object]] = {}
    for valores in hoja.iter_rows(min_row=2, values_only=True):
        clave = _limpiar_clave(valores[columnas["clave"] - 1])
        if not clave:
            continue
        registro_catalogo = {
            "Clave Catálogo": clave,
            "Nombre Catálogo": _valor_como_texto(valores[columnas["nombre"] - 1]),
            "CLABE": _valor_como_texto(
                valores[columnas["clabe"] - 1],
                quitar_apostrofe=True,
            ),
            "Banco Catálogo": _valor_como_texto(valores[columnas["banco"] - 1]),
            "ID Islas": valores[columnas["id_islas"] - 1],
            "ID Sindicato": valores[columnas["id_sindicato"] - 1],
        }
        catalogo_por_clave.setdefault(clave, registro_catalogo)
        cuenta_catalogo = _normalizar_cuenta(registro_catalogo["CLABE"])
        if cuenta_catalogo:
            catalogo_por_cuenta.setdefault(cuenta_catalogo, registro_catalogo)

    filas: list[dict[str, object]] = []
    incidencias: list[dict[str, object]] = []
    cuentas_verificadas = 0
    for registro in tabla_origen.to_dict("records"):
        clave_origen = str(registro["CLAVE"])
        cuenta_origen = _normalizar_cuenta(registro["No. Cuenta"])
        clave_temporal = _texto_normalizado(clave_origen) == "s n"
        if clave_temporal:
            encontrado = catalogo_por_cuenta.get(cuenta_origen, {})
        else:
            encontrado = catalogo_por_clave.get(clave_origen, {})
        fila = {
            **registro,
            "CLAVE": (
                encontrado.get("Clave Catálogo", clave_origen)
                if clave_temporal
                else clave_origen
            ),
            "REPT": "-",
            "Nombre Catálogo": encontrado.get("Nombre Catálogo"),
            "CLABE": encontrado.get("CLABE"),
            "Banco Catálogo": encontrado.get("Banco Catálogo"),
            "ID Islas": encontrado.get("ID Islas"),
            "ID Sindicato": encontrado.get("ID Sindicato"),
        }
        clabe_catalogo = _normalizar_cuenta(encontrado.get("CLABE"))
        cuenta_coincide = bool(encontrado) and cuenta_origen == clabe_catalogo
        fila.update(
            {
                "NOM": bool(encontrado)
                and _texto_igual_excel(registro["Nombre"], encontrado.get("Nombre Catálogo")),
                "CUEN": cuenta_coincide,
                "BAN": bool(encontrado)
                and _texto_igual_excel(registro["Banco"], encontrado.get("Banco Catálogo")),
            }
        )
        filas.append(fila)

        if cuenta_coincide:
            cuentas_verificadas += 1
        else:
            incidencias.append(
                {
                    "EMPRESA": registro["EMPRESA"],
                    "CLAVE": registro["CLAVE"],
                    "Nombre": registro["Nombre"],
                    "No. Cuenta": registro["No. Cuenta"],
                    "CLABE Catálogo": encontrado.get("CLABE"),
                    "Resultado": (
                        "Cuenta no coincide" if encontrado else "Clave no encontrada"
                    ),
                }
            )

    columnas_salida = [
        "EMPRESA",
        "BLOQUE",
        "CLAVE",
        "Nombre",
        "Total a Depositar",
        "No. Cuenta",
        "Banco",
        "REPT",
        "Nombre Catálogo",
        "CLABE",
        "Banco Catálogo",
        "ID Islas",
        "ID Sindicato",
        "NOM",
        "CUEN",
        "BAN",
    ]
    return (
        pd.DataFrame(filas, columns=columnas_salida),
        pd.DataFrame(incidencias),
        cuentas_verificadas,
    )


def obtener_no_coincidencias(tabla: pd.DataFrame) -> pd.DataFrame:
    comparaciones = {
        "NOM": ("Nombre", "Nombre Catálogo"),
        "CUEN": ("No. Cuenta", "CLABE"),
        "BAN": ("Banco", "Banco Catálogo"),
    }
    filas: list[dict[str, object]] = []
    for registro in tabla.to_dict("records"):
        for validacion, (columna_origen, columna_catalogo) in comparaciones.items():
            if bool(registro[validacion]):
                continue
            filas.append(
                {
                    "EMPRESA": registro["EMPRESA"],
                    "CLAVE": registro["CLAVE"],
                    "Nombre": registro["Nombre"],
                    "Validación": validacion,
                    "Valor del archivo": registro[columna_origen],
                    "Valor del Catálogo": registro[columna_catalogo],
                }
            )
    return pd.DataFrame(
        filas,
        columns=[
            "EMPRESA",
            "CLAVE",
            "Nombre",
            "Validación",
            "Valor del archivo",
            "Valor del Catálogo",
        ],
    )


def crear_tabla_corregida(tabla: pd.DataFrame) -> pd.DataFrame:
    corregida = tabla.copy()
    for indice, registro in corregida.iterrows():
        nombre_catalogo = registro["Nombre Catálogo"]
        clabe_catalogo = registro["CLABE"]
        banco_catalogo = registro["Banco Catálogo"]

        if pd.notna(nombre_catalogo) and _valor_como_texto(nombre_catalogo):
            corregida.at[indice, "Nombre"] = nombre_catalogo
        if pd.notna(clabe_catalogo) and _valor_como_texto(clabe_catalogo):
            corregida.at[indice, "No. Cuenta"] = clabe_catalogo
        if pd.notna(banco_catalogo) and _valor_como_texto(banco_catalogo):
            corregida.at[indice, "Banco"] = banco_catalogo

        corregida.at[indice, "NOM"] = _texto_igual_excel(
            corregida.at[indice, "Nombre"], nombre_catalogo
        )
        corregida.at[indice, "CUEN"] = _normalizar_cuenta(
            corregida.at[indice, "No. Cuenta"]
        ) == _normalizar_cuenta(clabe_catalogo)
        corregida.at[indice, "BAN"] = _texto_igual_excel(
            corregida.at[indice, "Banco"], banco_catalogo
        )
    return corregida


def filtrar_tabla(tabla: pd.DataFrame, busqueda: str) -> pd.DataFrame:
    termino = busqueda.strip()
    if not termino:
        return tabla
    patron = re.escape(termino)
    mascara = tabla.fillna("").astype(str).apply(
        lambda columna: columna.str.contains(patron, case=False, regex=True)
    ).any(axis=1)
    return tabla.loc[mascara]


def exportar_comparacion_excel(
    tabla: pd.DataFrame, contenido_catalogo: bytes, nombre_hoja: str
) -> bytes:
    libro_plantilla = load_workbook(
        BytesIO(contenido_catalogo), data_only=False, read_only=False
    )
    hojas_comparacion = [
        hoja
        for hoja in libro_plantilla.worksheets
        if _texto_normalizado(hoja.title) == "comparacion"
    ]
    if not hojas_comparacion:
        raise ValueError("El Catálogo no contiene la hoja 'Comparación' para tomar el formato.")
    plantilla = hojas_comparacion[0]

    libro_salida = Workbook()
    libro_salida.loaded_theme = libro_plantilla.loaded_theme
    hoja_salida = libro_salida.active
    hoja_salida.title = nombre_hoja[:31]

    encabezados_excel = [
        "EMPRESA",
        "BLOQUE",
        "CLAVE",
        "Nombre",
        "Total a Depositar",
        "No. Cuenta",
        "Banco",
        "REPT",
        "Nombre",
        "CLABE",
        "Banco",
        "ID Islas",
        "ID Sindicato",
        "NOM",
        "CUEN",
        "BAN",
    ]
    for numero_columna, encabezado in enumerate(encabezados_excel, 1):
        destino = hoja_salida.cell(1, numero_columna, encabezado)
        origen = plantilla.cell(1, numero_columna)
        destino.font = copy(origen.font)
        destino.fill = copy(origen.fill)
        destino.border = copy(origen.border)
        destino.alignment = copy(origen.alignment)
        destino.protection = copy(origen.protection)
        destino.number_format = origen.number_format

        letra = origen.column_letter
        dimension_origen = plantilla.column_dimensions[letra]
        dimension_destino = hoja_salida.column_dimensions[letra]
        dimension_destino.width = dimension_origen.width
        dimension_destino.hidden = dimension_origen.hidden

    for numero_fila, valores in enumerate(tabla.itertuples(index=False, name=None), 2):
        for numero_columna, valor in enumerate(valores, 1):
            if pd.isna(valor):
                valor = None
            elif hasattr(valor, "item"):
                valor = valor.item()
            destino = hoja_salida.cell(numero_fila, numero_columna, valor)
            origen = plantilla.cell(2, numero_columna)
            destino.font = copy(origen.font)
            destino.fill = copy(origen.fill)
            destino.border = copy(origen.border)
            destino.alignment = copy(origen.alignment)
            destino.protection = copy(origen.protection)
            destino.number_format = origen.number_format

    # Garantiza que las cuentas y CLABE se mantengan como texto y no pierdan ceros.
    for numero_fila in range(2, len(tabla) + 2):
        hoja_salida.cell(numero_fila, 6).number_format = "@"
        hoja_salida.cell(numero_fila, 10).number_format = "@"

    hoja_salida.row_dimensions[1].height = plantilla.row_dimensions[1].height
    hoja_salida.freeze_panes = "A2"
    hoja_salida.auto_filter.ref = f"A1:P{len(tabla) + 1}"
    hoja_salida.sheet_view.showGridLines = plantilla.sheet_view.showGridLines
    hoja_salida.sheet_properties.tabColor = copy(plantilla.sheet_properties.tabColor)

    salida = BytesIO()
    libro_salida.save(salida)
    return salida.getvalue()


def preparar_tabla_efectivo(tabla_efectivo: pd.DataFrame) -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Empresa": tabla_efectivo["EMPRESA"],
            "Clave": tabla_efectivo["CLAVE"],
            "Nombre": tabla_efectivo["Nombre"],
            "Importe": tabla_efectivo["Total a Depositar"],
            "CUENTA": "",
            "OBSERVACION": "",
        }
    )


def exportar_efectivo_excel(tabla: pd.DataFrame) -> bytes:
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Efectivo"

    total = float(tabla["Importe"].sum()) if not tabla.empty else 0.0
    hoja["D1"] = "EFECTIVO REQUERIDO"
    hoja["E1"] = total
    for celda in (hoja["D1"], hoja["E1"]):
        celda.font = Font(name="Arial", size=9, bold=True, color="FFC00000")
    hoja["D1"].alignment = Alignment(horizontal="center")
    hoja["E1"].number_format = '$#,##0.00'

    encabezados = ["Empresa", "Clave", "Nombre", "Importe", "CUENTA", "OBSERVACION"]
    relleno_encabezado = PatternFill("solid", fgColor="FF0D0D0D")
    for columna, encabezado in enumerate(encabezados, 2):
        celda = hoja.cell(2, columna, encabezado)
        celda.font = Font(name="Arial", size=8, bold=True, color="FFFFFFFF")
        celda.fill = relleno_encabezado
        celda.alignment = Alignment(horizontal="center")

    for numero_fila, valores in enumerate(tabla.itertuples(index=False, name=None), 3):
        for numero_columna, valor in enumerate(valores, 2):
            if pd.isna(valor):
                valor = None
            celda = hoja.cell(numero_fila, numero_columna, valor)
            celda.font = Font(name="Arial", size=8)
            if numero_columna == 2:
                celda.alignment = Alignment(horizontal="center")
            elif numero_columna in {4, 6}:
                celda.alignment = Alignment(horizontal="left")
        hoja.cell(numero_fila, 3).number_format = "@"
        hoja.cell(numero_fila, 5).number_format = '$#,##0.00'

    anchos = {"A": 13, "B": 13, "C": 13, "D": 38, "E": 13, "F": 13, "G": 18}
    for letra, ancho in anchos.items():
        hoja.column_dimensions[letra].width = ancho
    hoja.freeze_panes = "B3"
    hoja.auto_filter.ref = f"B2:G{len(tabla) + 2}"

    salida = BytesIO()
    libro.save(salida)
    return salida.getvalue()


def _catalogo_bancos(contenido: bytes) -> tuple[dict[str, str], dict[str, str]]:
    filas: list[tuple[object, ...]]
    if contenido.startswith(b"\xd0\xcf\x11\xe0"):
        import xlrd

        libro = xlrd.open_workbook(file_contents=contenido, on_demand=True)
        nombre_hoja = next(
            (
                nombre
                for nombre in libro.sheet_names()
                if _texto_normalizado(nombre) == "anexo catalogo de bancos"
            ),
            None,
        )
        if nombre_hoja is None:
            raise ValueError("No se encontró la hoja 'Anexo - Catálogo de Bancos'.")
        hoja = libro.sheet_by_name(nombre_hoja)
        filas = [tuple(hoja.row_values(fila)) for fila in range(hoja.nrows)]
    else:
        libro = load_workbook(BytesIO(contenido), data_only=True, read_only=True)
        hoja = next(
            (
                hoja
                for hoja in libro.worksheets
                if _texto_normalizado(hoja.title) == "anexo catalogo de bancos"
            ),
            None,
        )
        if hoja is None:
            raise ValueError("No se encontró la hoja 'Anexo - Catálogo de Bancos'.")
        filas = list(hoja.iter_rows(values_only=True))

    encabezado = None
    for indice, fila in enumerate(filas):
        normalizados = [_texto_normalizado(valor) for valor in fila]
        if "nombre de institucion" in normalizados and "clave transfer" in normalizados:
            encabezado = (indice, normalizados)
            break
    if encabezado is None:
        raise ValueError("No se encontraron NOMBRE DE INSTITUCION y CLAVE TRANSFER.")

    indice_fila, encabezados = encabezado
    columna_nombre = encabezados.index("nombre de institucion")
    columna_transfer = encabezados.index("clave transfer")
    columna_institucion = (
        encabezados.index("no inst") if "no inst" in encabezados else None
    )
    por_nombre: dict[str, str] = {}
    por_prefijo: dict[str, str] = {}
    for fila in filas[indice_fila + 1 :]:
        if max(columna_nombre, columna_transfer) >= len(fila):
            continue
        nombre = _texto_normalizado(fila[columna_nombre])
        clave_transfer = _valor_como_texto(fila[columna_transfer])
        if not nombre or not clave_transfer:
            continue
        por_nombre.setdefault(nombre, clave_transfer)
        if columna_institucion is not None and columna_institucion < len(fila):
            numero = re.sub(r"\D", "", _valor_como_texto(fila[columna_institucion]))
            if numero:
                por_prefijo.setdefault(numero[-3:].zfill(3), clave_transfer)
    return por_nombre, por_prefijo


def _nombre_banco_catalogo(nombre: object) -> str:
    normalizado = _texto_normalizado(nombre)
    alias = {
        "banbajio": "bajio",
        "bancomer": "bbva mexico",
        "bbva": "bbva mexico",
        "santader": "banco santander",
        "santander": "banco santander",
        "banorte": "banorte ixe",
    }
    return alias.get(normalizado, normalizado)


def _cuentas_csi_despues_del_corte(contenido: bytes) -> set[str]:
    libro = load_workbook(BytesIO(contenido), data_only=True, read_only=False)
    hoja = next(
        (hoja for hoja in libro.worksheets if _texto_normalizado(hoja.title) == "csi"),
        None,
    )
    if hoja is None:
        return set()
    estructura = _columnas_de_personas(hoja)
    if estructura is None:
        return set()
    fila_encabezado, columnas = estructura
    personas_vistas = False
    despues_del_corte = False
    cuentas: set[str] = set()

    for numero_fila in range(fila_encabezado + 1, hoja.max_row + 1):
        fila_vacia = all(
            hoja.cell(numero_fila, columna).value is None
            or str(hoja.cell(numero_fila, columna).value).strip() == ""
            for columna in range(1, hoja.max_column + 1)
        )
        if fila_vacia and personas_vistas and not despues_del_corte:
            despues_del_corte = True
            continue

        nombre = _valor_como_texto(hoja.cell(numero_fila, columnas["nombre"]).value)
        total = _decimal(hoja.cell(numero_fila, columnas["total"]).value)
        banco = _valor_como_texto(hoja.cell(numero_fila, columnas["banco"]).value)
        cuenta = _normalizar_cuenta(hoja.cell(numero_fila, columnas["cuenta"]).value)
        if not nombre or total is None or not banco or not cuenta:
            continue
        personas_vistas = True
        if despues_del_corte and _texto_normalizado(banco) != "efectivo":
            cuentas.add(cuenta)
    return cuentas


def crear_tabla_transferencias(
    tabla_corregida: pd.DataFrame,
    contenido_bancos: bytes,
    contenido_litografia: bytes,
) -> tuple[pd.DataFrame, list[str]]:
    bancos_por_nombre, bancos_por_prefijo = _catalogo_bancos(contenido_bancos)
    cuentas_con_recibo = _cuentas_csi_despues_del_corte(contenido_litografia)
    tabla_original, _ = extraer_tabla_comparacion(contenido_litografia)
    claves_originales = tabla_original["CLAVE"].tolist()
    filas: list[dict[str, object]] = []
    claves_para_recibos: list[str] = []
    bancos_sin_clave: set[str] = set()

    csi = tabla_corregida[tabla_corregida["EMPRESA"].map(_texto_normalizado) == "csi"]
    csi_no_santander = 0
    for _, registro in csi.iterrows():
        clabe = _normalizar_cuenta(registro["CLABE"])
        nombre_banco = _nombre_banco_catalogo(registro["Banco"])
        es_santander = nombre_banco == "banco santander" or clabe.startswith("014")
        if not es_santander:
            csi_no_santander += 1
    total_bloques = max(1, (csi_no_santander + 49) // 50)

    posicion_csi = 0
    for indice, registro in enumerate(tabla_corregida.to_dict("records")):
        empresa = _valor_como_texto(registro["EMPRESA"])
        clabe = _normalizar_cuenta(registro["CLABE"])
        nombre_banco = _nombre_banco_catalogo(registro["Banco"])
        es_santander = nombre_banco == "banco santander" or clabe.startswith("014")

        if es_santander:
            banco_transferencia = "SANTANDER"
        else:
            banco_transferencia = bancos_por_nombre.get(nombre_banco)
            if banco_transferencia is None and len(clabe) >= 3:
                banco_transferencia = bancos_por_prefijo.get(clabe[:3])
            if banco_transferencia is None:
                bancos_sin_clave.add(_valor_como_texto(registro["Banco"]))

        numero_bloque = ""
        if _texto_normalizado(empresa) == "csi":
            if es_santander:
                numero_bloque = "SANTANDER"
            else:
                posicion_csi += 1
                numero_bloque = f"{(posicion_csi - 1) // 50 + 1} DE {total_bloques}"

        recibo = ""
        if (
            _texto_normalizado(empresa) == "csi"
            and not es_santander
            and _normalizar_cuenta(registro["No. Cuenta"]) in cuentas_con_recibo
        ):
            recibo = "X"

        clave_original = (
            claves_originales[indice]
            if indice < len(claves_originales)
            else registro["CLAVE"]
        )
        claves_para_recibos.append(_limpiar_clave(clave_original))
        filas.append(
            {
                "CIA": empresa,
                "No Bloque": numero_bloque,
                "Clave": registro["CLAVE"],
                "Nombre": registro["Nombre"],
                "Banco": banco_transferencia or "",
                "CLABE": clabe,
                "CUENTA": clabe[6:-1] if es_santander and len(clabe) > 7 else "",
                "RECIBO": recibo,
                "Importe": registro["Total a Depositar"],
                "Com Banc": "",
            }
        )
    resultado = pd.DataFrame(filas)
    resultado.attrs["claves_originales"] = claves_para_recibos
    return resultado, sorted(bancos_sin_clave)


def exportar_transferencias_excel(tabla: pd.DataFrame) -> bytes:
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Transferencias"
    hoja["J1"] = float(tabla["Importe"].sum()) if not tabla.empty else 0.0
    hoja["J1"].font = Font(name="Arial", size=8, bold=True)
    hoja["J1"].number_format = '$#,##0.00'

    encabezados = list(tabla.columns)
    relleno = PatternFill("solid", fgColor="FF0D0D0D")
    for columna, encabezado in enumerate(encabezados, 2):
        celda = hoja.cell(2, columna, encabezado)
        celda.font = Font(name="Arial", size=8, bold=True, color="FFFFFFFF")
        celda.fill = relleno
        celda.alignment = Alignment(horizontal="center")

    for numero_fila, valores in enumerate(tabla.itertuples(index=False, name=None), 3):
        for numero_columna, valor in enumerate(valores, 2):
            if pd.isna(valor):
                valor = None
            celda = hoja.cell(numero_fila, numero_columna, valor)
            celda.font = Font(name="Arial", size=8)
        hoja.cell(numero_fila, 4).number_format = "@"
        hoja.cell(numero_fila, 7).number_format = "@"
        hoja.cell(numero_fila, 8).number_format = "@"
        hoja.cell(numero_fila, 10).number_format = '$#,##0.00'

    anchos = {
        "A": 3.55,
        "B": 7.44,
        "C": 11.44,
        "D": 9,
        "E": 41.89,
        "F": 11.11,
        "G": 19.11,
        "H": 12,
        "I": 11.55,
        "J": 12.33,
        "K": 11.55,
    }
    for letra, ancho in anchos.items():
        hoja.column_dimensions[letra].width = ancho
    hoja.freeze_panes = "B3"
    hoja.auto_filter.ref = f"B2:K{len(tabla) + 2}"

    salida = BytesIO()
    libro.save(salida)
    return salida.getvalue()


def preparar_interbancario_por_cia(tabla_transferencias: pd.DataFrame) -> pd.DataFrame:
    """Excluye Santander y conserva la información requerida por la macro bancaria."""
    mascara = ~tabla_transferencias["Banco"].map(
        lambda banco: "santander" in _texto_normalizado(banco)
    )
    return tabla_transferencias.loc[mascara].copy().reset_index(drop=True)


def _nombre_hoja_excel(nombre: object, usados: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "_", _valor_como_texto(nombre)).strip()
    base = (base or "SIN CIA")[:31]
    candidato = base
    numero = 2
    while candidato.casefold() in usados:
        sufijo = f"_{numero}"
        candidato = base[: 31 - len(sufijo)] + sufijo
        numero += 1
    usados.add(candidato.casefold())
    return candidato


def exportar_interbancario_por_cia(tabla_transferencias: pd.DataFrame) -> bytes:
    tabla = preparar_interbancario_por_cia(tabla_transferencias)
    if tabla.empty:
        raise ValueError("No hay transferencias interbancarias distintas de Santander.")

    libro = Workbook()
    libro.remove(libro.active)
    usados: set[str] = set()
    encabezados = [
        "CUENTA DE CARGO", "", "CUENTA DE ABONO", "BANCO RECEPTOR",
        "BENEFICIARIO", "SUCURSAL", "IMPORTE", "PLAZA BANXICO", "CONCEPTO",
        "", "EDO DE CUENTA", "RFC", "IVA", "Referencia Ordenante",
        "Forma de Aplicación",
    ]
    borde_encabezado = Border(
        left=Side(style="medium", color="FF000000"),
        right=Side(style="medium", color="FF000000"),
        top=Side(style="medium", color="FF000000"),
        bottom=Side(style="medium", color="FF000000"),
    )
    borde_datos = Border(
        left=Side(style="medium", color="FF000000"),
        right=Side(style="medium", color="FF000000"),
        bottom=Side(style="medium", color="FF000000"),
    )
    fecha_referencia = date.today().strftime("%d%m%Y")

    for cia, grupo in tabla.groupby("CIA", sort=False, dropna=False):
        hoja = libro.create_sheet(_nombre_hoja_excel(cia, usados))
        hoja.sheet_view.showGridLines = False

        for columna, encabezado in enumerate(encabezados, 1):
            celda = hoja.cell(7, columna, encabezado)
            celda.font = Font(name="Trebuchet MS", size=8, color="FFFFFFFF")
            celda.fill = PatternFill("solid", fgColor="FFFF0000")
            celda.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            celda.border = borde_encabezado
            if columna in {6, 8, 9}:
                celda.number_format = "@"

        for consecutivo, registro in enumerate(grupo.itertuples(index=False), 1):
            fila = consecutivo + 7
            valores = [
                "", "", registro.CLABE, registro.Banco, registro.Nombre, "",
                registro.Importe, "", str(consecutivo), "", "N",
                "XAXX010101000", "", fecha_referencia, 1,
            ]
            for columna, valor in enumerate(valores, 1):
                celda = hoja.cell(fila, columna, valor)
                celda.font = Font(name="Trebuchet MS", size=8)
                celda.alignment = Alignment(vertical="center")
                celda.border = borde_datos
            for columna in (3, 4, 5, 6, 8, 9, 11, 12, 13, 14):
                hoja.cell(fila, columna).number_format = "@"
            hoja.cell(fila, 7).number_format = '"$"#,##0.00'
            hoja.cell(fila, 15).alignment = Alignment(
                horizontal="center", vertical="center"
            )

        ultima_fila = len(grupo) + 7
        validacion_estado = DataValidation(
            type="list", formula1='"N,S"', allow_blank=False
        )
        validacion_estado.error = "Selecciona N o S."
        validacion_estado.errorTitle = "Valor no válido"
        validacion_estado.showErrorMessage = True
        validacion_forma = DataValidation(
            type="list", formula1='"1,2"', allow_blank=False
        )
        validacion_forma.error = "Selecciona 1 o 2."
        validacion_forma.errorTitle = "Valor no válido"
        validacion_forma.showErrorMessage = True
        hoja.add_data_validation(validacion_estado)
        hoja.add_data_validation(validacion_forma)
        validacion_estado.add(f"K8:K{ultima_fila}")
        validacion_forma.add(f"O8:O{ultima_fila}")

        anchos = {
            "A": 15.44, "B": 15.66, "C": 18, "D": 9.89, "E": 21.11,
            "F": 17.11, "G": 11.89, "H": 13.55, "I": 15, "J": 2,
            "K": 8.11, "L": 14.55, "M": 10.33, "N": 10.33, "O": 14.55,
        }
        for letra, ancho in anchos.items():
            hoja.column_dimensions[letra].width = ancho
        hoja.row_dimensions[7].height = 24
        for fila in range(8, ultima_fila + 1):
            hoja.row_dimensions[fila].height = 13.8
        hoja.freeze_panes = "A8"
        hoja.auto_filter.ref = f"A7:O{ultima_fila}"
        hoja.print_area = f"A7:O{ultima_fila}"
        hoja.page_setup.orientation = "landscape"
        hoja.page_setup.fitToWidth = 1
        hoja.page_setup.fitToHeight = 0
        hoja.sheet_properties.pageSetUpPr.fitToPage = True

    salida = BytesIO()
    libro.save(salida)
    return salida.getvalue()


def preparar_mismo_banco_por_cia(tabla_transferencias: pd.DataFrame) -> pd.DataFrame:
    tabla = tabla_transferencias.copy()
    es_santander = tabla.apply(
        lambda fila: (
            "santander" in _texto_normalizado(fila["Banco"])
            or _normalizar_cuenta(fila["CLABE"]).startswith("014")
        ),
        axis=1,
    )
    filas: list[dict[str, object]] = []
    for cia, grupo in tabla.groupby("CIA", sort=False, dropna=False):
        mascara_grupo = es_santander.loc[grupo.index]
        cantidad_interbancaria = int((~mascara_grupo).sum())
        santander = grupo.loc[mascara_grupo]
        for posicion, registro in enumerate(santander.to_dict("records"), 1):
            filas.append(
                {
                    "CIA": cia,
                    "CUENTA DE CARGO": "",
                    "CUENTA DE ABONO": registro["CUENTA"],
                    "IMPORTE": registro["Importe"],
                    "CONCEPTO": str(cantidad_interbancaria + posicion),
                    "FECHA": date.today().strftime("%d%m%Y"),
                }
            )
    return pd.DataFrame(
        filas,
        columns=[
            "CIA", "CUENTA DE CARGO", "CUENTA DE ABONO", "IMPORTE", "CONCEPTO", "FECHA"
        ],
    )


def exportar_mismo_banco_por_cia(tabla_transferencias: pd.DataFrame) -> bytes:
    tabla = preparar_mismo_banco_por_cia(tabla_transferencias)
    if tabla.empty:
        raise ValueError("No hay transferencias de Santander para generar Mismo Banco.")
    libro = Workbook()
    libro.remove(libro.active)
    usados: set[str] = set()
    encabezados = ["CUENTA DE CARGO", "CUENTA DE ABONO", "IMPORTE", "CONCEPTO", "FECHA"]
    borde_encabezado = Border(
        left=Side(style="medium", color="FF000000"),
        right=Side(style="medium", color="FF000000"),
        top=Side(style="medium", color="FF000000"),
        bottom=Side(style="medium", color="FF000000"),
    )
    borde_datos = Border(
        left=Side(style="medium", color="FF000000"),
        right=Side(style="medium", color="FF000000"),
        bottom=Side(style="medium", color="FF000000"),
    )
    for cia, grupo in tabla.groupby("CIA", sort=False, dropna=False):
        hoja = libro.create_sheet(_nombre_hoja_excel(cia, usados))
        hoja.sheet_view.showGridLines = False
        for columna, encabezado in enumerate(encabezados, 1):
            celda = hoja.cell(7, columna, encabezado)
            celda.font = Font(name="Trebuchet MS", size=8, color="FFFFFFFF")
            celda.fill = PatternFill("solid", fgColor="FFFF0000")
            celda.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            celda.border = borde_encabezado

        columnas_datos = encabezados
        for fila_excel, registro in enumerate(grupo.to_dict("records"), 8):
            for columna, nombre_columna in enumerate(columnas_datos, 1):
                valor = registro[nombre_columna]
                celda = hoja.cell(fila_excel, columna, None if valor == "" else valor)
                celda.font = Font(name="Trebuchet MS", size=8)
                celda.alignment = Alignment(vertical="center")
                celda.border = borde_datos
            hoja.cell(fila_excel, 1).number_format = "@"
            hoja.cell(fila_excel, 2).number_format = "@"
            hoja.cell(fila_excel, 3).number_format = '"$"#,##0.00'
            hoja.cell(fila_excel, 4).number_format = "@"
            hoja.cell(fila_excel, 5).number_format = "@"

        ultima_fila = len(grupo) + 7
        for letra, ancho in {
            "A": 21.89, "B": 19.66, "C": 24.66, "D": 47, "E": 22.66,
        }.items():
            hoja.column_dimensions[letra].width = ancho
        hoja.row_dimensions[7].height = 24
        for fila in range(8, ultima_fila + 1):
            hoja.row_dimensions[fila].height = 13.8
        hoja.freeze_panes = "A8"
        hoja.auto_filter.ref = f"A7:E{ultima_fila}"
        hoja.print_area = f"A7:E{ultima_fila}"
        hoja.page_setup.orientation = "landscape"
        hoja.page_setup.fitToWidth = 1
        hoja.page_setup.fitToHeight = 0
        hoja.sheet_properties.pageSetUpPr.fitToPage = True

    salida = BytesIO()
    libro.save(salida)
    return salida.getvalue()


def preparar_detalle(contenido: bytes) -> tuple[pd.DataFrame, list[float]]:
    libro = load_workbook(BytesIO(contenido), data_only=True, read_only=False)
    hoja_resumen = next(
        (
            hoja
            for hoja in libro.worksheets
            if _texto_normalizado(hoja.title) == "resumen"
        ),
        None,
    )
    if hoja_resumen is None:
        raise ValueError("El archivo no contiene la hoja Resumen.")
    importes = [
        float(fila.importe.valor)
        for fila in _filas_del_resumen(hoja_resumen)
        if fila.importe.valor is not None
    ]
    if not importes:
        raise ValueError("No se encontraron importes en la columna Suma de Resumen.")
    if len(importes) > 7:
        raise ValueError("Detalle admite como máximo siete importes (filas 4 a 10).")

    filas = []
    for importe in importes:
        comision = importe * 0.03
        subtotal = importe + comision
        iva = subtotal * 0.16
        total = subtotal + iva
        filas.append(
            {
                "Empresa": "",
                "Importe": importe,
                "Comisión": comision,
                "Subtotal": subtotal,
                "IVA": iva,
                "Total": total,
                "Dev 8 Puntos": (iva / 16) * 8,
                "ADEUDO": total,
                "TOTAL DEPÓSITO": 0.0,
            }
        )
    return pd.DataFrame(filas), importes


def exportar_detalle_excel(importes: list[float], semana: int | None = None) -> bytes:
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Detalle"
    semana = semana or date.today().isocalendar().week
    hoja["B2"] = f"Nómina Extra semana {semana}"
    hoja["I2"] = "SOLO SE PIDE FONDEO SI PAGAN EL TOTAL"

    grupos = [
        ("M2:Q2", "Primer Depósito"),
        ("R2:V2", "Segundo Depósito"),
        ("W2:AA2", "Tercer Depósito"),
        ("AB2:AF2", "Cuarto Depósito"),
    ]
    for rango, titulo in grupos:
        hoja.merge_cells(rango)
        celda = hoja[rango.split(":")[0]]
        celda.value = titulo
        celda.fill = PatternFill("solid", fgColor="FFE5F4F7")
        celda.font = Font(name="Calibri", size=9)
        celda.alignment = Alignment(horizontal="center", vertical="center")

    encabezados = {
        "B3": "Empresa",
        "C3": "Importe",
        "D3": "Comisión",
        "E3": "Subtotal",
        "F3": "IVA",
        "G3": "Total",
        "I3": "Dev 8 Puntos",
        "K3": "ADEUDO",
        "L3": "TOTAL DEPÓSITO",
    }
    encabezados_deposito = [
        "Importe",
        "Fecha de depósito",
        "No Op Bancaria",
        "Factura",
        "Fecha de factura",
    ]
    for inicio in (13, 18, 23, 28):
        for desplazamiento, titulo in enumerate(encabezados_deposito):
            encabezados[f"{hoja.cell(3, inicio + desplazamiento).column_letter}3"] = titulo

    borde = Border(
        left=Side(style="thin", color="FFB7C9D6"),
        right=Side(style="thin", color="FFB7C9D6"),
        top=Side(style="thin", color="FFB7C9D6"),
        bottom=Side(style="thin", color="FFB7C9D6"),
    )
    for referencia, titulo in encabezados.items():
        celda = hoja[referencia]
        celda.value = titulo
        celda.font = Font(name="Calibri", size=9, bold=True)
        celda.fill = PatternFill("solid", fgColor="FFDDEBF7")
        celda.alignment = Alignment(horizontal="center", vertical="center")
        celda.border = borde
    hoja["I3"].fill = PatternFill("solid", fgColor="FF000000")
    hoja["I3"].font = Font(name="Calibri", size=9, bold=True, color="FFFFFFFF")
    hoja["K3"].fill = PatternFill("solid", fgColor="FFC6E0B4")
    hoja["K3"].font = Font(name="Calibri", size=9, bold=True, color="FFC00000")
    for columna in range(12, 33):
        hoja.cell(3, columna).fill = PatternFill("solid", fgColor="FFE5F4F7")

    hoja["B2"].font = Font(name="Calibri", size=11, bold=True)
    hoja["I2"].font = Font(name="Calibri", size=8, bold=True)
    hoja["I2"].fill = PatternFill("solid", fgColor="FFFFFF00")
    hoja["I2"].alignment = Alignment(vertical="center")

    for indice, importe in enumerate(importes):
        hoja.cell(4 + indice, 3, importe)

    for fila in range(4, 11):
        hoja[f"D{fila}"] = f"=+C{fila}*0.03"
        hoja[f"E{fila}"] = f"=+C{fila}+D{fila}"
        hoja[f"F{fila}"] = f"=+E{fila}*0.16"
        hoja[f"G{fila}"] = f"=+E{fila}+F{fila}"
        hoja[f"I{fila}"] = f"=+(F{fila}/16)*8"
        hoja[f"K{fila}"] = f"=G{fila}-L{fila}"
        hoja[f"L{fila}"] = f"=+M{fila}+R{fila}+W{fila}+AB{fila}"
        for columna in range(2, 33):
            hoja.cell(fila, columna).border = borde

    formulas_total = {
        "B11": "SUMA",
        "C11": "=SUM(C4:C9)",
        "D11": "=SUM(D4:D9)",
        "E11": "=SUM(E4:E9)",
        "F11": "=SUM(F4:F9)",
        "G11": "=SUM(G4:G9)",
        "I11": "=SUM(I4:I10)",
        "K11": "=SUM(K4:K9)",
        "L11": "=SUM(L4:L9)",
        "M11": "=SUM(M4:M9)",
        "R11": "=SUM(R4:R8)",
        "W11": "=SUM(W4:W8)",
        "AB11": "=SUM(AB4:AB8)",
    }
    for referencia, valor in formulas_total.items():
        celda = hoja[referencia]
        celda.value = valor
        celda.font = Font(name="Calibri", size=9, bold=True, color="FFFFFFFF")
        celda.fill = PatternFill("solid", fgColor="FF000000")
        celda.border = borde
        celda.alignment = Alignment(vertical="center")
    hoja["I11"].fill = PatternFill("solid", fgColor="FFFFFF99")
    hoja["I11"].font = Font(name="Calibri", size=8, bold=True)

    formato_moneda = '"$"#,##0.00;[Red]\-"$"#,##0.00'
    for fila in range(4, 12):
        for columna in (3, 4, 5, 6, 7, 9, 11, 12, 13, 18, 23, 28):
            hoja.cell(fila, columna).number_format = formato_moneda
    for fila in range(4, 11):
        for columna in (14, 17, 19, 22, 24, 27, 29, 32):
            hoja.cell(fila, columna).number_format = "dd/mm/yyyy"

    anchos = {
        "B": 22.22, "C": 8.89, "D": 13, "E": 10, "F": 8.89, "G": 10.11,
        "H": 1.89, "I": 8.89, "J": 2, "K": 10.44, "L": 12.33,
    }
    for letra, ancho in anchos.items():
        hoja.column_dimensions[letra].width = ancho
    for columna in range(13, 33):
        hoja.column_dimensions[hoja.cell(1, columna).column_letter].width = 13
    hoja.freeze_panes = "B4"
    hoja.auto_filter.ref = "B3:AF11"
    libro.calculation.fullCalcOnLoad = True
    libro.calculation.forceFullCalc = True
    libro.calculation.calcMode = "auto"

    salida = BytesIO()
    libro.save(salida)
    return salida.getvalue()


def preparar_listado_recibos(tabla_transferencias: pd.DataFrame) -> pd.DataFrame:
    """Conserva las transferencias para las que debe entregarse un recibo."""
    recibo = tabla_transferencias["RECIBO"].fillna("").astype(str).str.strip().str.upper()
    mascara = recibo != "X"
    filas = tabla_transferencias.loc[mascara]
    claves_guardadas = tabla_transferencias.attrs.get("claves_originales")
    if claves_guardadas is not None and len(claves_guardadas) == len(tabla_transferencias):
        claves = pd.Series(claves_guardadas, index=tabla_transferencias.index).loc[mascara]
    else:
        claves = filas["Clave"].map(_limpiar_clave)
    return pd.DataFrame(
        {
            "Empresa": filas["CIA"],
            "Clave": claves,
            "ID": "",
            "Nombre del Socio": filas["Nombre"],
            "Importe": filas["Importe"],
        }
    ).reset_index(drop=True)


def _numero_entero_en_letra(numero: int, apocopar: bool = True) -> str:
    unidades = [
        "CERO", "UNO", "DOS", "TRES", "CUATRO", "CINCO", "SEIS", "SIETE",
        "OCHO", "NUEVE", "DIEZ", "ONCE", "DOCE", "TRECE", "CATORCE", "QUINCE",
        "DIECISEIS", "DIECISIETE", "DIECIOCHO", "DIECINUEVE", "VEINTE",
    ]
    centenas = {
        2: "DOSCIENTOS", 3: "TRESCIENTOS", 4: "CUATROCIENTOS",
        5: "QUINIENTOS", 6: "SEISCIENTOS", 7: "SETECIENTOS",
        8: "OCHOCIENTOS", 9: "NOVECIENTOS",
    }

    if numero < 0:
        return "MENOS " + _numero_entero_en_letra(abs(numero), apocopar)
    if numero <= 20:
        texto = unidades[numero]
    elif numero < 30:
        texto = "VEINTI" + unidades[numero - 20].lower().upper()
    elif numero < 100:
        decenas = {
            3: "TREINTA", 4: "CUARENTA", 5: "CINCUENTA", 6: "SESENTA",
            7: "SETENTA", 8: "OCHENTA", 9: "NOVENTA",
        }
        texto = decenas[numero // 10]
        if numero % 10:
            texto += " Y " + unidades[numero % 10]
    elif numero == 100:
        texto = "CIEN"
    elif numero < 1000:
        texto = "CIENTO" if numero < 200 else centenas[numero // 100]
        if numero % 100:
            texto += " " + _numero_entero_en_letra(numero % 100, apocopar)
    elif numero < 1_000_000:
        miles, resto = divmod(numero, 1000)
        # PesosMN utiliza "UN MIL", por eso se conserva esa misma redacción.
        texto = _numero_entero_en_letra(miles, True) + " MIL"
        if resto:
            texto += " " + _numero_entero_en_letra(resto, apocopar)
    elif numero < 1_000_000_000:
        millones, resto = divmod(numero, 1_000_000)
        texto = (
            "UN MILLON"
            if millones == 1
            else _numero_entero_en_letra(millones, False) + " MILLONES"
        )
        if resto:
            texto += " " + _numero_entero_en_letra(resto, apocopar)
    else:
        raise ValueError("El importe es demasiado grande para convertirlo a letra.")

    if apocopar:
        if texto.endswith("VEINTIUNO"):
            texto = texto[:-9] + "VEINTIUN"
        elif texto.endswith(" Y UNO"):
            texto = texto[:-6] + " Y UN"
        elif texto.endswith(" UNO"):
            texto = texto[:-4] + " UN"
        elif texto == "UNO":
            texto = "UN"
    return texto


def importe_en_letra(valor: object) -> str:
    decimal = _decimal(valor) or Decimal("0")
    decimal = decimal.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    entero = int(decimal)
    centavos = int((decimal - entero) * 100)
    return f"( {_numero_entero_en_letra(entero)} PESOS {centavos:02d}/100 M.N.)"


def preparar_para_imprimir(tabla_recibos: pd.DataFrame, semana: int) -> pd.DataFrame:
    consecutivos = range(1, len(tabla_recibos) + 1)
    return pd.DataFrame(
        {
            "NOI": consecutivos,
            "NOMBRE": tabla_recibos["Nombre del Socio"],
            "IMPORTE": tabla_recibos["Importe"],
            "OPERACIÓN": str(semana),
            "CLAVE": tabla_recibos["Clave"],
            "Folio": range(1, len(tabla_recibos) + 1),
            "Cantidad": tabla_recibos["Importe"].map(importe_en_letra),
            "C.C": "",
        }
    )


def _celda_word(
    texto: object,
    ancho: int,
    *,
    alineacion: str = "left",
    negrita: bool = False,
    tamano: int = 22,
    columnas: int = 1,
) -> str:
    contenido = escape(str(texto if texto is not None else ""))
    expansion = f'<w:gridSpan w:val="{columnas}"/>' if columnas > 1 else ""
    negrita_xml = "<w:b/>" if negrita else ""
    bordes = "".join(
        f'<w:{lado} w:val="single" w:sz="10" w:space="0" w:color="000000"/>'
        for lado in ("top", "left", "bottom", "right", "insideH", "insideV")
    )
    return (
        f'<w:tc><w:tcPr><w:tcW w:w="{ancho}" w:type="dxa"/>{expansion}'
        f'<w:tcBorders>{bordes}</w:tcBorders></w:tcPr>'
        f'<w:p><w:pPr><w:jc w:val="{alineacion}"/></w:pPr>'
        f'<w:r><w:rPr>{negrita_xml}<w:rFonts w:ascii="Arial" w:hAnsi="Arial"/>'
        f'<w:sz w:val="{tamano}"/><w:szCs w:val="{tamano}"/></w:rPr>'
        f'<w:t xml:space="preserve">{contenido}</w:t></w:r></w:p></w:tc>'
    )


def _tabla_word(filas: list[list[str]], anchos: list[int]) -> str:
    cuadricula = "".join(f'<w:gridCol w:w="{ancho}"/>' for ancho in anchos)
    filas_xml = "".join(f"<w:tr>{''.join(fila)}</w:tr>" for fila in filas)
    return (
        '<w:tbl><w:tblPr><w:tblW w:w="0" w:type="auto"/>'
        '<w:tblLayout w:type="fixed"/></w:tblPr>'
        f"<w:tblGrid>{cuadricula}</w:tblGrid>{filas_xml}</w:tbl>"
    )


def _pagina_comprobante_word(
    folio: int,
    semana: int,
    nombre: object,
    clave: object,
    importe: object,
) -> str:
    tabla_titulo = _tabla_word(
        [[
            _celda_word(
                "COMPROBANTE DE PAGO", 6506, alineacion="center", negrita=True, tamano=36
            ),
            _celda_word("", 1134),
            _celda_word(folio, 1276, alineacion="center", negrita=True, tamano=28),
        ]],
        [6506, 1134, 1276],
    )
    tabla_datos = _tabla_word(
        [
            [_celda_word(semana, 8926, alineacion="center", negrita=True, columnas=2)],
            [
                _celda_word("Nombre:", 2093, negrita=True),
                _celda_word(nombre, 6833, negrita=True),
            ],
            [
                _celda_word("Clave:", 2093, negrita=True),
                _celda_word(_limpiar_clave(clave), 6833, negrita=True),
            ],
        ],
        [2093, 6833],
    )
    importe_decimal = float(_decimal(importe) or Decimal("0"))
    tabla_importe = _tabla_word(
        [
            [
                _celda_word("IMPORTE LETRA:", 6516, alineacion="center", negrita=True),
                _celda_word("IMPORTE", 2410, alineacion="center", negrita=True),
            ],
            [
                _celda_word(
                    importe_en_letra(importe_decimal), 6516, alineacion="center", tamano=20
                ),
                _celda_word(
                    f"$    {importe_decimal:,.2f}", 2410, alineacion="center", tamano=24
                ),
            ],
            [_celda_word("", 6516), _celda_word("", 2410)],
            [_celda_word("", 6516), _celda_word("FIRMA", 2410, alineacion="center", negrita=True)],
        ],
        [6516, 2410],
    )
    espacio = '<w:p><w:pPr><w:spacing w:after="120"/></w:pPr></w:p>'
    return tabla_titulo + espacio + tabla_datos + espacio + tabla_importe


def exportar_comprobantes_word(tabla_recibos: pd.DataFrame, semana: int) -> bytes:
    paginas: list[str] = []
    for folio, registro in enumerate(tabla_recibos.to_dict("records"), 1):
        if paginas:
            paginas.append('<w:p><w:r><w:br w:type="page"/></w:r></w:p>')
        paginas.append(
            _pagina_comprobante_word(
                folio,
                semana,
                registro["Nombre del Socio"],
                registro["Clave"],
                registro["Importe"],
            )
        )

    documento = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        f'<w:body>{"".join(paginas)}'
        '<w:sectPr><w:pgSz w:w="12242" w:h="15842"/>'
        '<w:pgMar w:top="1701" w:right="284" w:bottom="1701" w:left="1418" '
        'w:header="709" w:footer="709" w:gutter="0"/></w:sectPr>'
        '</w:body></w:document>'
    )
    tipos = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/word/document.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
        '</Types>'
    )
    relaciones = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="word/document.xml"/></Relationships>'
    )
    relaciones_documento = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"/>'
    )
    salida = BytesIO()
    with ZipFile(salida, "w", ZIP_DEFLATED) as archivo:
        archivo.writestr("[Content_Types].xml", tipos)
        archivo.writestr("_rels/.rels", relaciones)
        archivo.writestr("word/document.xml", documento)
        archivo.writestr("word/_rels/document.xml.rels", relaciones_documento)
    return salida.getvalue()


def exportar_paquete_litografia_zip(
    excel_general: bytes,
    excel_interbancario: bytes,
    excel_mismo_banco: bytes,
    word_comprobantes: bytes,
    semana: int,
    semana_mismo_banco: int,
    fecha_archivo: str,
) -> bytes:
    salida = BytesIO()
    with ZipFile(salida, "w", ZIP_DEFLATED) as archivo:
        archivo.writestr(
            f"litografia_general_semana_{semana}_{fecha_archivo}.xlsx",
            excel_general,
        )
        archivo.writestr(
            f"interbancario_por_cia_{fecha_archivo}.xlsx",
            excel_interbancario,
        )
        archivo.writestr(
            f"mismo_banco_semana_{semana_mismo_banco}_{fecha_archivo}.xlsx",
            excel_mismo_banco,
        )
        archivo.writestr(
            f"comprobantes_pago_semana_{semana}_{fecha_archivo}.docx",
            word_comprobantes,
        )
    return salida.getvalue()


def _formato_acuse(hoja) -> None:
    hoja.sheet_view.showGridLines = False
    for letra, ancho in {
        "A": 13, "B": 15, "C": 15, "D": 18, "E": 16, "F": 16, "G": 3, "H": 3
    }.items():
        hoja.column_dimensions[letra].width = ancho
    hoja.page_setup.orientation = "portrait"
    hoja.page_setup.fitToWidth = 1
    hoja.page_setup.fitToHeight = 1
    hoja.sheet_properties.pageSetUpPr.fitToPage = True
    hoja.page_margins.left = 0.35
    hoja.page_margins.right = 0.35
    hoja.page_margins.top = 0.45
    hoja.page_margins.bottom = 0.45


def _encabezado_acuse(hoja, fila_empresa: int, fila_fecha: int) -> None:
    hoja.merge_cells(start_row=fila_empresa, start_column=1, end_row=fila_empresa, end_column=6)
    empresa = hoja.cell(fila_empresa, 1, "Islas Gower y Compañía Sucesores S. en C. de C.V.")
    empresa.font = Font(name="Arial", size=12, bold=True)
    empresa.alignment = Alignment(horizontal="center")
    hoja.merge_cells(start_row=fila_fecha, start_column=4, end_row=fila_fecha, end_column=6)
    celda_fecha = hoja.cell(fila_fecha, 4, date.today())
    celda_fecha.font = Font(name="Arial", size=10)
    celda_fecha.alignment = Alignment(horizontal="right")
    celda_fecha.number_format = '[$-es-MX]d "de" mmmm "de" yyyy'


def _firma_acuse(hoja, fila: int) -> None:
    hoja.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=6)
    hoja.cell(fila, 1, "Atentamente").alignment = Alignment(horizontal="center")
    hoja.cell(fila, 1).font = Font(name="Arial", size=10)
    hoja.merge_cells(start_row=fila + 1, start_column=1, end_row=fila + 1, end_column=6)
    hoja.cell(fila + 1, 1, "Lexa Consulting").alignment = Alignment(horizontal="center")
    hoja.cell(fila + 1, 1).font = Font(name="Arial", size=10, bold=True)


def _crear_hoja_efectivo_nomina(
    libro: Workbook, total_efectivo: float, semana: int
) -> None:
    hoja = libro.active
    hoja.title = "EFECTIVO NOMINA"
    _formato_acuse(hoja)
    moneda = '$#,##0.00'

    _encabezado_acuse(hoja, 1, 2)
    hoja.merge_cells("A4:D4")
    hoja["A4"] = "A quien corresponda:"
    hoja.merge_cells("A6:F7")
    hoja["A6"] = (
        "Por este medio informo la cantidad solicitada para entrega "
        f"correspondiente a la semana {semana}"
    )
    hoja["A6"].alignment = Alignment(wrap_text=True, vertical="center")
    hoja.merge_cells("B9:C9")
    hoja["B9"] = "NÓMINA EFECTIVO (CH)"
    hoja["D9"] = total_efectivo
    hoja["D9"].number_format = moneda
    hoja["D12"] = "=SUM(D9:D11)"
    hoja["D12"].number_format = moneda
    hoja["D12"].font = Font(name="Arial", size=10, bold=True)
    hoja["A13"] = "Agradezco firme al calce de recibido"
    _firma_acuse(hoja, 15)

    _encabezado_acuse(hoja, 26, 27)
    hoja.merge_cells("A29:D29")
    hoja["A29"] = "A quien corresponda:"
    hoja.merge_cells("A31:F32")
    hoja["A31"] = hoja["A6"].value
    hoja["A31"].alignment = Alignment(wrap_text=True, vertical="center")
    hoja.merge_cells("B34:C34")
    hoja["B34"] = "NÓMINA EFECTIVO (CH)"
    hoja["D34"] = total_efectivo
    hoja["D34"].number_format = moneda
    hoja["D37"] = "=SUM(D34:D36)"
    hoja["D37"].number_format = moneda
    hoja["D37"].font = Font(name="Arial", size=10, bold=True)
    hoja["A38"] = "Agradezco firme al calce de recibido"
    _firma_acuse(hoja, 40)
    hoja.print_area = "A1:F41"


def _crear_hoja_dev_puntos(
    libro: Workbook, total_dev_puntos: float, semana: int
) -> None:
    hoja = libro.create_sheet("DEV PUNTOS")
    _formato_acuse(hoja)
    moneda = '$#,##0.00'

    _encabezado_acuse(hoja, 1, 2)
    hoja.merge_cells("A4:D4")
    hoja["A4"] = "A quien corresponda:"
    hoja.merge_cells("A6:F7")
    hoja["A6"] = (
        "Por este medio informo la cantidad que se entrega correspondiente a la "
        f"devolución de puntos de la Semana {semana}"
    )
    hoja["A6"].alignment = Alignment(wrap_text=True, vertical="center")
    hoja["D10"] = total_dev_puntos
    hoja["D10"].number_format = moneda
    hoja["D13"] = "=SUM(D9:D12)"
    hoja["D13"].number_format = moneda
    hoja["D13"].font = Font(name="Arial", size=10, bold=True)
    hoja["A17"] = "Agradezco firme al calce de recibido"
    _firma_acuse(hoja, 19)

    _encabezado_acuse(hoja, 29, 30)
    hoja.merge_cells("A32:D32")
    hoja["A32"] = "A quien corresponda:"
    hoja.merge_cells("A34:F35")
    hoja["A34"] = hoja["A6"].value
    hoja["A34"].alignment = Alignment(wrap_text=True, vertical="center")
    hoja["D39"] = total_dev_puntos
    hoja["D39"].number_format = moneda
    hoja["D41"] = "=SUM(D36:D40)"
    hoja["D41"].number_format = moneda
    hoja["D41"].font = Font(name="Arial", size=10, bold=True)
    hoja["A45"] = "Agradezco firme al calce de recibido"
    _firma_acuse(hoja, 47)
    hoja.print_area = "A1:F48"


def _crear_hoja_listado_recibos(
    libro: Workbook, tabla: pd.DataFrame, semana: int
) -> None:
    hoja = libro.create_sheet("LISTADO RECIBOS")
    hoja.sheet_view.showGridLines = False
    hoja.merge_cells("A1:F1")
    hoja["A1"] = "Islas Gower y Compañía Sucesores S. en C. de C.V."
    hoja["A1"].font = Font(name="Arial", size=12, bold=True)
    hoja["A1"].alignment = Alignment(horizontal="center")
    hoja.merge_cells("E2:F2")
    hoja["E2"] = date.today()
    hoja["E2"].number_format = '[$-es-MX]d "de" mmmm "de" yyyy'
    hoja["E2"].alignment = Alignment(horizontal="right")
    hoja["A3"] = "C.P. Gabriel Brenes Velázquez"
    hoja["A4"] = "Litografía Magno Graf"
    hoja["A5"] = (
        "Relación de recibos entregados, correspondientes a la nómina extra "
        f"de la semana {semana}"
    )

    encabezados = list(tabla.columns)
    borde = Border(bottom=Side(style="thin", color="FF000000"))
    for columna, encabezado in enumerate(encabezados, 1):
        celda = hoja.cell(6, columna, encabezado)
        celda.font = Font(name="Arial", size=9, bold=True)
        celda.fill = PatternFill("solid", fgColor="FFD9EAF7")
        celda.alignment = Alignment(horizontal="center")
        celda.border = borde

    for fila_excel, valores in enumerate(tabla.itertuples(index=False, name=None), 7):
        for columna, valor in enumerate(valores, 1):
            if pd.isna(valor):
                valor = None
            celda = hoja.cell(fila_excel, columna, valor)
            celda.font = Font(name="Arial", size=8)
        hoja.cell(fila_excel, 2).number_format = "@"
        hoja.cell(fila_excel, 5).number_format = '$#,##0.00'

    ultima_fila = 6 + len(tabla)
    fila_conteo = ultima_fila + 2
    hoja.cell(fila_conteo, 4, "Recibos:").font = Font(name="Arial", size=9, bold=True)
    hoja.cell(fila_conteo, 5, f"=SUBTOTAL(2,E7:E{ultima_fila})")
    hoja.cell(fila_conteo, 5).font = Font(name="Arial", size=9, bold=True)
    _firma_acuse(hoja, fila_conteo + 2)

    for letra, ancho in {"A": 12, "B": 11, "C": 8, "D": 42, "E": 14, "F": 13}.items():
        hoja.column_dimensions[letra].width = ancho
    hoja.freeze_panes = "A7"
    hoja.auto_filter.ref = f"A6:E{ultima_fila}"
    hoja.print_title_rows = "1:6"
    hoja.print_area = f"A1:F{fila_conteo + 3}"
    hoja.page_setup.orientation = "portrait"
    hoja.page_setup.fitToWidth = 1
    hoja.page_setup.fitToHeight = 0
    hoja.sheet_properties.pageSetUpPr.fitToPage = True


def _crear_hoja_para_imprimir(libro: Workbook, tabla: pd.DataFrame) -> None:
    hoja = libro.create_sheet("Para imprimir por C.C")
    hoja.sheet_view.showGridLines = True
    encabezados = list(tabla.columns)
    relleno = PatternFill("solid", fgColor="FFC00000")
    for columna, encabezado in enumerate(encabezados, 1):
        celda = hoja.cell(1, columna, encabezado)
        celda.font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
        celda.fill = relleno
        celda.alignment = Alignment(horizontal="center")

    for fila_excel, valores in enumerate(tabla.itertuples(index=False, name=None), 2):
        for columna, valor in enumerate(valores, 1):
            if pd.isna(valor):
                valor = None
            celda = hoja.cell(fila_excel, columna, valor)
            celda.font = Font(name="Calibri", size=11)
        hoja.cell(fila_excel, 1).alignment = Alignment(horizontal="right")
        hoja.cell(fila_excel, 3).number_format = (
            '_-"$"* #,##0.00_-;\\-"$"* #,##0.00_-;_-"$"* "-"??_-;_-@_-'
        )
        hoja.cell(fila_excel, 4).number_format = "@"
        hoja.cell(fila_excel, 5).number_format = "@"
        hoja.cell(fila_excel, 6).alignment = Alignment(horizontal="right")

    anchos = {
        "A": 10.89, "B": 35.11, "C": 14.44, "D": 27.55,
        "E": 11.11, "F": 11.55, "G": 58.11, "H": 32.89,
    }
    for letra, ancho in anchos.items():
        hoja.column_dimensions[letra].width = ancho
    hoja.row_dimensions[1].height = 21
    ultima_fila = len(tabla) + 1
    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = f"A1:H{ultima_fila}"
    hoja.print_area = f"A1:H{ultima_fila}"
    hoja.page_setup.orientation = "landscape"
    hoja.page_setup.fitToWidth = 1
    hoja.page_setup.fitToHeight = 0
    hoja.sheet_properties.pageSetUpPr.fitToPage = True


def exportar_acuses_excel(
    total_efectivo: float,
    total_dev_puntos: float,
    tabla_recibos: pd.DataFrame,
    semana: int | None = None,
) -> bytes:
    semana = semana or date.today().isocalendar().week
    libro = Workbook()
    _crear_hoja_efectivo_nomina(libro, total_efectivo, semana)
    _crear_hoja_dev_puntos(libro, total_dev_puntos, semana)
    _crear_hoja_listado_recibos(libro, tabla_recibos, semana)
    _crear_hoja_para_imprimir(libro, preparar_para_imprimir(tabla_recibos, semana))
    libro.calculation.fullCalcOnLoad = True
    libro.calculation.forceFullCalc = True
    libro.calculation.calcMode = "auto"
    salida = BytesIO()
    libro.save(salida)
    return salida.getvalue()


def _copiar_hoja_excel(hoja_origen, hoja_destino) -> None:
    for fila in hoja_origen.iter_rows():
        for origen in fila:
            if isinstance(origen, MergedCell):
                continue
            destino = hoja_destino.cell(origen.row, origen.column, origen.value)
            if origen.has_style:
                destino.font = copy(origen.font)
                destino.fill = copy(origen.fill)
                destino.border = copy(origen.border)
                destino.alignment = copy(origen.alignment)
                destino.protection = copy(origen.protection)
                destino.number_format = origen.number_format

    for rango in hoja_origen.merged_cells.ranges:
        hoja_destino.merge_cells(str(rango))
    for letra, dimension in hoja_origen.column_dimensions.items():
        destino = hoja_destino.column_dimensions[letra]
        destino.width = dimension.width
        destino.hidden = dimension.hidden
        destino.bestFit = dimension.bestFit
    for numero, dimension in hoja_origen.row_dimensions.items():
        destino = hoja_destino.row_dimensions[numero]
        destino.height = dimension.height
        destino.hidden = dimension.hidden

    hoja_destino.freeze_panes = hoja_origen.freeze_panes
    hoja_destino.auto_filter.ref = hoja_origen.auto_filter.ref
    hoja_destino.sheet_view.showGridLines = hoja_origen.sheet_view.showGridLines
    hoja_destino.sheet_format.defaultColWidth = hoja_origen.sheet_format.defaultColWidth
    hoja_destino.sheet_format.defaultRowHeight = hoja_origen.sheet_format.defaultRowHeight
    hoja_destino.page_margins = copy(hoja_origen.page_margins)
    hoja_destino.page_setup = copy(hoja_origen.page_setup)
    hoja_destino.print_options = copy(hoja_origen.print_options)
    hoja_destino.sheet_properties.tabColor = copy(hoja_origen.sheet_properties.tabColor)
    if hoja_origen.print_area:
        hoja_destino.print_area = hoja_origen.print_area
    if hoja_origen.print_title_rows:
        hoja_destino.print_title_rows = hoja_origen.print_title_rows
    if hoja_origen.print_title_cols:
        hoja_destino.print_title_cols = hoja_origen.print_title_cols


def exportar_descarga_general(
    excel_detalle: bytes,
    excel_comparacion_corregida: bytes,
    excel_cuentas_nuevas: bytes,
    excel_efectivo: bytes,
    excel_transferencias: bytes,
    excel_acuses: bytes,
) -> bytes:
    libro_salida = load_workbook(BytesIO(excel_detalle), data_only=False)
    fuentes = [
        (excel_comparacion_corregida, "Comparación corregida"),
        (excel_cuentas_nuevas, "Cuentas nuevas"),
        (excel_efectivo, "Efectivo"),
        (excel_transferencias, "Transferencias"),
    ]
    for contenido, nombre in fuentes:
        libro_origen = load_workbook(BytesIO(contenido), data_only=False)
        hoja_origen = libro_origen.active
        if nombre in libro_salida.sheetnames:
            del libro_salida[nombre]
        hoja_destino = libro_salida.create_sheet(nombre[:31])
        _copiar_hoja_excel(hoja_origen, hoja_destino)

    libro_acuses = load_workbook(BytesIO(excel_acuses), data_only=False)
    for hoja_origen in libro_acuses.worksheets:
        nombre = hoja_origen.title[:31]
        if nombre in libro_salida.sheetnames:
            del libro_salida[nombre]
        hoja_destino = libro_salida.create_sheet(nombre)
        _copiar_hoja_excel(hoja_origen, hoja_destino)

    libro_salida.calculation.fullCalcOnLoad = True
    libro_salida.calculation.forceFullCalc = True
    libro_salida.calculation.calcMode = "auto"
    salida = BytesIO()
    libro_salida.save(salida)
    return salida.getvalue()


def vista_litografia() -> None:
    fecha_descarga = date.today().strftime("%d%m%Y")
    st.header("Litografía")
    st.write(
        "Carga el libro de Excel para comparar la **Suma** indicada en la hoja "
        "**Resumen** contra la **Suma** de cada una de las demás hojas."
    )

    archivo = st.file_uploader(
        "Cargar archivo de Litografía",
        type=["xlsx", "xlsm"],
        key="excel_litografia",
        help="Formatos admitidos: .xlsx y .xlsm",
    )
    if archivo is None:
        st.info("Carga el archivo para ejecutar la primera verificación.")
        return

    try:
        contenido_litografia = archivo.getvalue()
        resultados, nombre_resumen = analizar_excel(contenido_litografia)
        tabla_detalle, importes_detalle = preparar_detalle(contenido_litografia)
    except Exception as error:
        st.error(f"No se pudo analizar el archivo: {error}")
        return

    coinciden = int((resultados["Estado"] == "Coincide").sum())
    no_coinciden = int((resultados["Estado"] == "No coincide").sum())
    faltantes = int((resultados["Estado"] == "Falta dato").sum())
    columna_1, columna_2, columna_3 = st.columns(3)
    columna_1.metric("Coinciden", coinciden)
    columna_2.metric("No coinciden", no_coinciden)
    columna_3.metric("Faltan datos", faltantes)

    st.subheader("Primera verificación")
    st.dataframe(
        resultados,
        hide_index=True,
        use_container_width=True,
        column_config={
            "Suma en Resumen": st.column_config.NumberColumn(format="$ %.2f"),
            "Suma en hoja": st.column_config.NumberColumn(format="$ %.2f"),
            "Diferencia": st.column_config.NumberColumn(format="$ %.2f"),
        },
    )

    if no_coinciden:
        st.warning(f"Hay {no_coinciden} hoja(s) cuya suma no coincide con Resumen.")
    elif faltantes:
        st.warning(
            "No fue posible localizar todos los importes. Revisa las referencias de celda; "
            "si las sumas son fórmulas sin resultado guardado, abre y guarda el archivo en Excel."
        )
    else:
        st.success("Todas las sumas coinciden con el Resumen.")

    st.divider()
    st.subheader("Base de datos de personas")
    archivo_catalogo = st.file_uploader(
        "Cargar Catálogo de personas",
        type=["xlsx", "xlsm"],
        key="catalogo_litografia",
        help="El archivo debe contener una hoja llamada Catálogo.",
    )
    if archivo_catalogo is None:
        st.info("Carga el Catálogo para generar la tabla de comparación.")
        return

    try:
        contenido_catalogo = archivo_catalogo.getvalue()
        personas_catalogo = validar_catalogo(contenido_catalogo)
        tabla_origen, tabla_efectivo = extraer_tabla_comparacion(archivo.getvalue())
        tabla, _incidencias_cuenta, _cuentas_verificadas = crear_tabla_con_catalogo(
            tabla_origen, contenido_catalogo
        )
        no_coincidencias = obtener_no_coincidencias(tabla)
        tabla_cuentas_nuevas = detectar_cuentas_nuevas(
            tabla_origen, contenido_catalogo
        )
        tabla_corregida = crear_tabla_corregida(tabla)
        excel_comparacion = exportar_comparacion_excel(
            tabla, contenido_catalogo, "Comparación"
        )
        excel_corregido = exportar_comparacion_excel(
            tabla_corregida, contenido_catalogo, "Comparación corregida"
        )
        excel_cuentas_nuevas = exportar_cuentas_nuevas_excel(
            tabla_cuentas_nuevas
        )
        tabla_efectivo_preparada = preparar_tabla_efectivo(tabla_efectivo)
        excel_efectivo = exportar_efectivo_excel(tabla_efectivo_preparada)
    except Exception as error:
        st.error(f"No se pudo preparar la comparación: {error}")
        return

    st.success(f"Catálogo cargado correctamente: {personas_catalogo:,} registros.")
    st.subheader("Tabla para comparación")
    st.caption(
        "La clave no contiene puntos ni espacios. En Nombre sólo se quitaron "
        "espacios al principio y al final. Las columnas de Catálogo se obtienen "
        "por CLAVE. NOM, CUEN y BAN validan Nombre, Cuenta y Banco, respectivamente."
    )
    metrica_nom, metrica_cuen, metrica_ban = st.columns(3)
    metrica_nom.metric("NOM correctos", int(tabla["NOM"].sum()))
    metrica_cuen.metric("CUEN correctas", int(tabla["CUEN"].sum()))
    metrica_ban.metric("BAN correctos", int(tabla["BAN"].sum()))

    busqueda_original = st.text_input(
        "Buscar en la tabla de comparación",
        placeholder="Empresa, clave, nombre, cuenta, banco...",
        key="buscar_tabla_comparacion_litografia",
    )
    tabla_visible = filtrar_tabla(tabla, busqueda_original)
    if busqueda_original.strip():
        st.caption(f"Se encontraron {len(tabla_visible)} registro(s).")

    st.dataframe(
        tabla_visible,
        hide_index=True,
        use_container_width=True,
        column_config={
            "Total a Depositar": st.column_config.NumberColumn(format="$ %.2f"),
            "No. Cuenta": st.column_config.TextColumn(),
            "Nombre Catálogo": st.column_config.TextColumn("Nombre"),
            "CLABE": st.column_config.TextColumn(),
            "Banco Catálogo": st.column_config.TextColumn("Banco"),
        },
    )
    st.download_button(
        "Descargar tabla de comparación",
        data=excel_comparacion,
        file_name=f"comparacion_litografia_{fecha_descarga}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="descargar_comparacion_litografia",
    )
    st.subheader("No coincidencias")
    if no_coincidencias.empty:
        st.success("Nombre, cuenta y banco coinciden en todos los registros.")
    else:
        st.warning(f"Se encontraron {len(no_coincidencias)} validación(es) sin coincidencia.")
        st.dataframe(no_coincidencias, hide_index=True, use_container_width=True)

    st.subheader("Cuentas nuevas")
    st.caption(
        "Incluye registros cuya clave original o CLABE no existe en el Catálogo. "
        "Las claves S/N se excluyen de esta validación."
    )
    if tabla_cuentas_nuevas.empty:
        st.success("No se detectaron cuentas nuevas.")
    else:
        st.warning(f"Se detectaron {len(tabla_cuentas_nuevas)} cuenta(s) nueva(s).")
        busqueda_cuentas_nuevas = st.text_input(
            "Buscar en cuentas nuevas",
            placeholder="Empresa, clave, nombre, CLABE, motivo...",
            key="buscar_cuentas_nuevas_litografia",
        )
        cuentas_nuevas_visibles = filtrar_tabla(
            tabla_cuentas_nuevas, busqueda_cuentas_nuevas
        )
        st.dataframe(
            cuentas_nuevas_visibles,
            hide_index=True,
            use_container_width=True,
            column_config={
                "Clave": st.column_config.TextColumn(),
                "CLABE interbancaria": st.column_config.TextColumn(),
                "Importe": st.column_config.NumberColumn(format="$ %.2f"),
            },
        )
    st.download_button(
        "Descargar cuentas nuevas",
        data=excel_cuentas_nuevas,
        file_name=f"cuentas_nuevas_{fecha_descarga}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="descargar_cuentas_nuevas_litografia",
    )

    st.subheader("Tabla de comparación corregida")
    st.caption(
        "Nombre, No. Cuenta y Banco se sustituyen por los valores del Catálogo. "
        "Los datos originales sólo se muestran en la tabla anterior y no se modifica "
        "ninguno de los archivos cargados."
    )
    busqueda_corregida = st.text_input(
        "Buscar en la tabla corregida",
        placeholder="Empresa, clave, nombre, cuenta, banco...",
        key="buscar_tabla_corregida_litografia",
    )
    tabla_corregida_visible = filtrar_tabla(tabla_corregida, busqueda_corregida)
    if busqueda_corregida.strip():
        st.caption(f"Se encontraron {len(tabla_corregida_visible)} registro(s).")
    st.dataframe(
        tabla_corregida_visible,
        hide_index=True,
        use_container_width=True,
        column_config={
            "Total a Depositar": st.column_config.NumberColumn(format="$ %.2f"),
            "No. Cuenta": st.column_config.TextColumn(),
            "Nombre Catálogo": st.column_config.TextColumn("Nombre"),
            "CLABE": st.column_config.TextColumn(),
            "Banco Catálogo": st.column_config.TextColumn("Banco"),
        },
    )
    st.download_button(
        "Descargar tabla corregida",
        data=excel_corregido,
        file_name=f"comparacion_litografia_corregida_{fecha_descarga}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="descargar_comparacion_litografia_corregida",
    )
    st.divider()
    st.subheader("Tabla de efectivo")
    if tabla_efectivo_preparada.empty:
        st.info("No se encontraron registros con Banco igual a Efectivo.")
    else:
        efectivo_requerido = float(tabla_efectivo_preparada["Importe"].sum())
        st.metric("EFECTIVO REQUERIDO", f"$ {efectivo_requerido:,.2f}")
        st.caption(
            "Incluye todos los registros cuyo Banco es Efectivo, incluso filas sin "
            "clave como EFECTIVO *** CSIP ****."
        )
        busqueda_efectivo = st.text_input(
            "Buscar en la tabla de efectivo",
            placeholder="Empresa, clave, nombre, importe...",
            key="buscar_tabla_efectivo_litografia",
        )
        efectivo_visible = filtrar_tabla(tabla_efectivo_preparada, busqueda_efectivo)
        if busqueda_efectivo.strip():
            st.caption(f"Se encontraron {len(efectivo_visible)} registro(s).")
        st.dataframe(
            efectivo_visible,
            hide_index=True,
            use_container_width=True,
            column_config={
                "Importe": st.column_config.NumberColumn(format="$ %.2f"),
                "Clave": st.column_config.TextColumn(),
            },
        )
        st.download_button(
            "Descargar tabla de efectivo",
            data=excel_efectivo,
            file_name=f"efectivo_litografia_{fecha_descarga}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="descargar_efectivo_litografia",
        )

    st.divider()
    st.subheader("Tabla de transferencias")
    archivo_bancos = st.file_uploader(
        "Cargar Catálogo de Bancos",
        type=["xls", "xlsx", "xlsm"],
        key="catalogo_bancos_litografia",
        help="Debe contener la hoja Anexo - Catálogo de Bancos.",
    )
    if archivo_bancos is None:
        st.info("Carga el Catálogo de Bancos para generar las transferencias.")
        return

    try:
        tabla_transferencias, bancos_sin_clave = crear_tabla_transferencias(
            tabla_corregida,
            archivo_bancos.getvalue(),
            archivo.getvalue(),
        )
        excel_transferencias = exportar_transferencias_excel(tabla_transferencias)
        excel_interbancario = exportar_interbancario_por_cia(tabla_transferencias)
        excel_mismo_banco = exportar_mismo_banco_por_cia(tabla_transferencias)
        semana_mismo_banco = (date.today() - timedelta(days=7)).isocalendar().week
        tabla_recibos = preparar_listado_recibos(tabla_transferencias)
        total_efectivo_acuse = float(tabla_efectivo_preparada["Importe"].sum())
        total_dev_puntos = float(tabla_detalle["Dev 8 Puntos"].sum())
    except Exception as error:
        st.error(f"No se pudo preparar la tabla de transferencias: {error}")
        return

    total_transferencias = float(tabla_transferencias["Importe"].sum())
    bloques_csi = tabla_transferencias.loc[
        (tabla_transferencias["CIA"].map(_texto_normalizado) == "csi")
        & tabla_transferencias["No Bloque"].str.contains(" DE ", na=False),
        "No Bloque",
    ].nunique()
    metrica_transferencias_1, metrica_transferencias_2, metrica_transferencias_3 = st.columns(3)
    metrica_transferencias_1.metric("Transferencias", len(tabla_transferencias))
    metrica_transferencias_2.metric("Total", f"$ {total_transferencias:,.2f}")
    metrica_transferencias_3.metric("Bloques CSI", bloques_csi)

    if bancos_sin_clave:
        st.warning(
            "No se encontró CLAVE TRANSFER para: " + ", ".join(bancos_sin_clave)
        )

    busqueda_transferencias = st.text_input(
        "Buscar en la tabla de transferencias",
        placeholder="Empresa, bloque, clave, nombre, banco, CLABE...",
        key="buscar_tabla_transferencias_litografia",
    )
    transferencias_visibles = filtrar_tabla(
        tabla_transferencias, busqueda_transferencias
    )
    if busqueda_transferencias.strip():
        st.caption(f"Se encontraron {len(transferencias_visibles)} registro(s).")
    st.dataframe(
        transferencias_visibles,
        hide_index=True,
        use_container_width=True,
        column_config={
            "Clave": st.column_config.TextColumn(),
            "CLABE": st.column_config.TextColumn(),
            "CUENTA": st.column_config.TextColumn(),
            "Importe": st.column_config.NumberColumn(format="$ %.2f"),
        },
    )
    st.download_button(
        "Descargar tabla de transferencias",
        data=excel_transferencias,
        file_name=f"transferencias_litografia_{fecha_descarga}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="descargar_transferencias_litografia",
    )

    semana_actual = date.today().isocalendar().week
    semana_guardada = str(st.session_state.get("semana_acuses_litografia", "")).strip()
    semana_valida = True
    if not semana_guardada:
        semana_seleccionada = semana_actual
    else:
        try:
            semana_seleccionada = int(semana_guardada)
            if not 1 <= semana_seleccionada <= 53:
                raise ValueError
        except ValueError:
            semana_seleccionada = semana_actual
            semana_valida = False

    try:
        excel_detalle = exportar_detalle_excel(importes_detalle, semana_seleccionada)
        excel_acuses = exportar_acuses_excel(
            total_efectivo_acuse,
            total_dev_puntos,
            tabla_recibos,
            semana_seleccionada,
        )
        word_comprobantes = exportar_comprobantes_word(
            tabla_recibos, semana_seleccionada
        )
        excel_general = exportar_descarga_general(
            excel_detalle,
            excel_corregido,
            excel_cuentas_nuevas,
            excel_efectivo,
            excel_transferencias,
            excel_acuses,
        )
        paquete_litografia = exportar_paquete_litografia_zip(
            excel_general,
            excel_interbancario,
            excel_mismo_banco,
            word_comprobantes,
            semana_seleccionada,
            semana_mismo_banco,
            fecha_descarga,
        )
    except Exception as error:
        st.error(f"No se pudieron generar los archivos de descarga: {error}")
        return

    st.divider()
    st.subheader(f"Detalle — Nómina Extra semana {semana_seleccionada}")
    st.caption(
        "Los importes conservan el orden de la columna Suma de Resumen. "
        "Empresa queda vacía y el Excel descargado incluye las fórmulas de B2:AF11."
    )
    busqueda_detalle = st.text_input(
        "Buscar en la tabla de detalle",
        placeholder="Importe, comisión, subtotal...",
        key="buscar_tabla_detalle_litografia",
    )
    detalle_visible = filtrar_tabla(tabla_detalle, busqueda_detalle)
    if busqueda_detalle.strip():
        st.caption(f"Se encontraron {len(detalle_visible)} registro(s).")
    st.dataframe(
        detalle_visible,
        hide_index=True,
        use_container_width=True,
        column_config={
            columna: st.column_config.NumberColumn(format="$ %.2f")
            for columna in tabla_detalle.columns
            if columna != "Empresa"
        },
    )
    st.download_button(
        "Descargar tabla de detalle",
        data=excel_detalle,
        file_name=f"detalle_semana_{semana_seleccionada}_{fecha_descarga}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="descargar_detalle_litografia",
    )

    st.divider()
    st.subheader(f"Acuses y recibos — Semana {semana_seleccionada}")
    st.caption(
        "Incluye EFECTIVO NOMINA, DEV PUNTOS, LISTADO RECIBOS y Para imprimir por C.C. "
        "Ambos listados contienen solamente las transferencias que no tienen X en RECIBO. "
        "Las claves se toman del primer Excel y se eliminan sus puntos y espacios."
    )
    metrica_acuse_1, metrica_acuse_2, metrica_acuse_3 = st.columns(3)
    metrica_acuse_1.metric("Efectivo nómina", f"$ {total_efectivo_acuse:,.2f}")
    metrica_acuse_2.metric("Dev puntos", f"$ {total_dev_puntos:,.2f}")
    metrica_acuse_3.metric("Recibos", len(tabla_recibos))
    busqueda_recibos = st.text_input(
        "Buscar en el listado de recibos",
        placeholder="Empresa, clave, nombre o importe...",
        key="buscar_listado_recibos_litografia",
    )
    recibos_visibles = filtrar_tabla(tabla_recibos, busqueda_recibos)
    if busqueda_recibos.strip():
        st.caption(f"Se encontraron {len(recibos_visibles)} registro(s).")
    st.dataframe(
        recibos_visibles,
        hide_index=True,
        use_container_width=True,
        column_config={
            "Clave": st.column_config.TextColumn(),
            "ID": st.column_config.TextColumn(),
            "Importe": st.column_config.NumberColumn(format="$ %.2f"),
        },
    )
    columna_semana, columna_descarga_acuses = st.columns([1, 2])
    with columna_semana:
        st.text_input(
            "Semana",
            placeholder=f"Actual: {semana_actual}",
            key="semana_acuses_litografia",
            help="Déjala vacía para utilizar automáticamente la semana actual.",
        )
    with columna_descarga_acuses:
        st.caption(f"Todos los documentos se generarán con la semana {semana_seleccionada}.")
        st.download_button(
            "Descargar acuses y listado de recibos",
            data=excel_acuses,
            file_name=f"acuse_recibos_semana_{semana_seleccionada}_{fecha_descarga}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="descargar_acuses_litografia",
        )
        st.download_button(
            "Descargar comprobantes de pago (Word)",
            data=word_comprobantes,
            file_name=f"comprobantes_pago_semana_{semana_seleccionada}_{fecha_descarga}.docx",
            mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            key="descargar_comprobantes_word_litografia",
        )
    if not semana_valida:
        st.warning(
            f"La semana debe ser un número del 1 al 53. Mientras tanto se usará la semana actual ({semana_actual})."
        )

    st.subheader("Descarga general")
    st.caption(
        "Un solo libro con Detalle, Comparación corregida, Cuentas nuevas, Efectivo, "
        "Transferencias, EFECTIVO NOMINA, DEV PUNTOS, LISTADO RECIBOS y Para imprimir por C.C."
    )
    columna_general, columna_interbancario, columna_mismo_banco, columna_paquete = st.columns(4)
    with columna_general:
        st.download_button(
            "Descargar archivo general de Litografía",
            data=excel_general,
            file_name=f"litografia_general_semana_{semana_seleccionada}_{fecha_descarga}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="descargar_litografia_general",
            use_container_width=True,
        )
    with columna_interbancario:
        st.download_button(
            "Descargar interbancario por CIA",
            data=excel_interbancario,
            file_name=f"interbancario_por_cia_{fecha_descarga}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="descargar_interbancario_por_cia_litografia",
            use_container_width=True,
        )
    with columna_mismo_banco:
        st.download_button(
            f"Mismo banco (semana {semana_mismo_banco})",
            data=excel_mismo_banco,
            file_name=f"mismo_banco_semana_{semana_mismo_banco}_{fecha_descarga}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="descargar_mismo_banco_por_cia_litografia",
            use_container_width=True,
        )
    with columna_paquete:
        st.download_button(
            "Descargar todo (.zip)",
            data=paquete_litografia,
            file_name=f"paquete_litografia_semana_{semana_seleccionada}_{fecha_descarga}.zip",
            mime="application/zip",
            key="descargar_paquete_litografia",
            use_container_width=True,
        )
